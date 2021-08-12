#! /usr/bin/python3
# -*-coding:utf-8-*-
# by lgh
# @Time     :2019/3/20 9:07
# @Author   :LGH
# @Site     :
# @File     :batch_modification0.1.py
# @Software :

import requests
import json
import time
import pymysql
import os
import openpyxl
import re
import traceback
import multiprocessing
import datetime
from collections import defaultdict


class MyDb:
    def __init__(self):
        self.name = 'channelaction'
        self.passwd = 'd0177951d1493f49'
        self.dbname = 'cmdb'
        self.port = '3306'
        self.host = '38.83.75.10'
        self.db = None

    def connect_db(self):
        self.db = pymysql.connect(host=self.host, user=self.name, password=self.passwd, database=self.dbname)
        return self.db

    def get_info(self, name):
        while True:
            try:
                print(name)
                cursor = self.db.cursor()
                select_command = u"select hostIp from cdnChannel where channelMark='%s';" % name
                cursor.execute(select_command)
                results = cursor.fetchall()
                # print(type(results))
                name_list = []
                # print('results: {}'.format(results))
                if len(results) != 0:
                    name_list.append(name)
                    server_dict1 = defaultdict(list)
                    for n in results:
                        server_dict1[n[0]].append(name)
                else:
                    select_command1 = 'select hostIp,channelMark from cdnChannel where channelMark like "%'+name+'@%";'
                    # print(select_command1)
                    cursor.execute(select_command1)
                    results = cursor.fetchall()
                    print('results: {}'.format(results))
                    server_dict1 = defaultdict(list)
                    for n in results:
                        if n[1] not in name_list:
                            name_list.append(n[1])
                        server_dict1[n[0]].append(n[1])
                    # print('get-server_dict: {}'.format(server_dict1))
                    # print('name_list: {}'.format(name_list))
                # 返回ip列表
                break
            except pymysql.OperationalError as e:
                print(e)
                self.db.ping()
                self.connect_db()
        print('name_list:{}'.format(name_list))
        print('server_dict:{}'.format(server_dict1))
        return name_list, server_dict1

    def close(self, connect_db):
        connect_db.close()


class Cms:
    def __init__(self, ip):
        self.ip = ip
        self.cookie = ''
        self.rows = {}

    def login(self):
        # get cookies
        ip = self.ip
        url = 'http://%s:8088/user/login' % ip
        # print('登录：%s' % url)
        # re = requests.get(url)
        # print(re.headers)
        # # print(type(re.headers['Set-Cookie']))
        # jsession_id = re.headers['Set-cookie']
        # print('cookie:%s' % (re.headers))
        host = '%s:8088' % ip
        # print(host)
        referer = 'http://%s:8088/user/login' % ip
        # print(referer)
        # # cookie = jsession_id + ";cms_language=en_US; loginUser=%E5%88%98%E5%86%A0%E5%8D%8E; loginPass=zVFaXk37"
        # self.cookie = jsession_id
        # print('cookie:%s' % self.cookie)
        headers = {
            # 'Host': '178.132.6.62:8088',
            'Host': host,
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/78.0.3904.108 Safari/537.36',
            'Accept': "application/json",
            'Accept-Language': 'zh,en-US;q=0.7,en;q=0.3',
            'Accept-Encoding': 'gzip, deflate',
            'Content-Type': 'application/json; charset=utf-8',
            'Referer': referer,
            'Connection': 'keep - alive',
            # 'DNT': '1',
            'origin': 'http://%s:8088' % ip,
            # 'Cookie': self.cookie
        }

        login_url = 'http://%s:8088/cdn/submitLogin' % ip
        filepath = os.getcwd()+r'/login.txt'
        with open(filepath, 'r') as f:
            list = f.readlines()
        name = list[0].strip()
        passwd = list[1].strip()
        # name = 'admin'
        # passwd = 'Q5t0wRLN'
        login_post = {

            "userName": name,
            "password": passwd,
            "type": "account",
            "local": "en_US",
        }
        # print(login_post)
        # 注意此时用的json参数
        res = requests.post(url=login_url, json=login_post, headers=headers)

        # SHIROSESSION = 97229d9a-912b-4a19-a685-9cef7a3fddd4;Path = /;
        # print(res.headers['Set-Cookie'])
        cookie = res.headers['Set-Cookie']
        # cookie = 'SHIROSESSION=97229d9a-912b-4a19-a685-9cef7a3fddd4;Path = /'
        # 正向后视断定
        # test = re.findall(r'(?<=SHIROSESSION=)[^;]+', cookie)
        self.cookie = re.findall(r'[^;]+', cookie)[0]
        # print(test)
        # print(self.cookie)
        # print(res.status_code)
        # print(res.text)
        # print(res.headers)
        # print(res.headers['date'])

    def select_ch(self, cpcode):

        headers = {
            'Host': '%s:8088' % self.ip,
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0',
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'Accept-Language': 'zh,en-US;q=0.7,en;q=0.3',
            'Accept-Encoding': 'gzip, deflate',
            'Referer': 'http://%s:8088/live/recording' % self.ip,
            # 'Content-Type': 'application/json; charset=utf-8',
            # 'X-Requested-With': 'XMLHttpRequest',
            # 'Content-Length': '113',
            # 'DNT': '1',
            'Connection': 'keep-alive',
            'Cookie': self.cookie
        }

        # select_post = {
        #     'pageSize': '100',
        #     'channel_id': cpcode,
        # }
        select_url = 'http://'+self.ip+':8088/cdn/record?channel_id='+cpcode+'&pageSize='+'100'
        # res = requests.post(url=select_url, data=select_post, headers=headers)
        while True:
            try:
                res = requests.get(url=select_url, headers=headers)
                break
            except Exception as e:
                print(res.text)
                print(e)
        # print(res.text)
        # print(type(res.text))
        # res.text输出内容为字符串，需要转换为字典
        result = json.loads(res.text)
        # print(result)
        # id = result['list'][0]['channel_id']
        # 获取id
        # post_data = {'id': id}
        # select_url1 = 'http://%s:8180/cms//servermonitor/getChannelRecordById.action' % self.ip
        #
        # ress = requests.post(url=select_url1, data=post_data, headers=headers)
        # result1 = json.loads(ress.text)
        # 提取唯一标识
        # print(result)
        if result['pagination']['total'] == 1:
            cpcode = result['list'][0]['channel_id']
            result1 = result['list'][0]
            self.rows = result1
            # print(self.rows)
            print('唯一标识：'+self.rows['channel_id']+'\n'+'节目名称：'+self.rows['channel_name']+'\n'+'来源分组：'+
                  self.rows['origin']+'\n')
        else:
            for ch_dict in result['list']:
                if ch_dict['channel_id'] == cpcode:
                    self.rows = ch_dict
        # print(self.rows)
        # channel_name = result1['channel_name']
        # if result1['state'] == '100':
        #     stream_status = '正常'
        # elif result1['state'] == '101':
        #     stream_status = '中断'
        # else:
        #     stream_status = '已下发'
        # if result1['issue_state'] == '105':
        #     status = '未下发'
        # else:
        #     status = '已下发'
        #
        # origin = result1['origin']
        # source_url = result1['source_url']
        # print(('频道名称：%s\n' % channel_name)+('频道唯一标识：%s\n' % cpcode)+('流状态：%s\n' % stream_status)
        #       + ('下发状态：%s\n' % status)+('源链接%s\n' % source_url)+('正在使用的链接：%s\n' % self.rows['currenturl'])
        #       + ('备注：%s\n' % origin) + ("\n"))
        # print('频道名称：%s' % channel_name)
        # print('频道唯一标识：%s' % cpcode)
        # print('流状态：%s' % stream_status)
        # print('下发状态：%s' % status)
        # print('源链接%s' % source_url):
        # print('正在使用的链接：%s' % self.rows['currenturl'])
        # print('备注：%s' % origin)
        # print("\n")

    def updata_ch(self, tv_name=None, url=None, origin=None):
        headers = {
            'Host': '%s:8081' % self.ip,
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0',
            'Accept': 'application/json',
            'Accept-Language': 'zh-CN',
            'Accept-Encoding': 'gzip, deflate',
            'Referer': 'http://%s:8088/live/recording' % self.ip,
            'Content-Type': 'application/json; charset=utf-8',
            'X-Requested-With': 'XMLHttpRequest',
            # 'Content-Length': '113',
            # 'DNT': '1',
            'Connection': 'keep-alive',
            'Cookie': self.cookie
        }
        if tv_name is not None:
            self.rows['channel_name'] = tv_name
            # print(cpcode+'更新名称')
        if url is not None:
            self.rows['source_url'] = url
            # print(cpcode+'更新源url')
        if origin is not None:
            self.rows['origin'] = origin
            # print(cpcode+'更新来源分组')
        # print(self.rows)
        updata_post = {
            'channel_id': self.rows['channel_id'],
            'channel_name':self.rows['channel_name'],
            'id': self.rows['id'],
            'issue_state': self.rows['issue_state'],
            'on_demand': self.rows['on_demand'],
            'origin': self.rows['origin'],
            'select_app': self.rows['select_app'],
            'source_url': self.rows['source_url'],
            'store_time': self.rows['store_time']
        }
        # print('updata_post:' + str(updata_post))
        update_url = 'http://%s:8088/cdn/record' % self.ip

        res = requests.post(url=update_url, json=updata_post, headers=headers)
        # print(res.text)
        # status = 'false'
        # if res.text == 'true':
        #     q.put(('------------------------------')+('%s上%s更新成功\n' % (self.ip, self.rows['cpcontentid']))
        #           +('------------------------------\n'))
        #     # print('%s上%s更新成功' % (self.ip, self.rows['cpcontentid']))
        #     # print('------------------------------')
        # else:
        #     q.put(('------------------------------')+('%s上%s更新失败\n' % (self.ip, self.rows['cpcontentid']))
        #           +(res.text)+('\n------------------------------\n'))
        #     # print('%s上%s更新失败' % (self.ip, self.rows['cpcontentid']))
        #     # print(res.text)
        #     # print('------------------------------')
        time.sleep(1)
        # print(res.text)
        # {"code": 200, "message": null}
        res_status = json.loads(res.text).get('message', "null")
        # print(res_status)
        return res_status


def get_ch():
    tv_info = {}
    file_path = os.getcwd()+'/Modify_information.xlsx'
    print("表格路径为：{}".format(file_path))
    wb = openpyxl.load_workbook(file_path)
    sh = wb.active
    rows = sh.max_row
    print("获取到{}行数据".format(rows))
    for row in range(rows):
        tv_info[sh.cell(row=row+1, column=1).value] = sh.cell(row=row+1, column=2).value
    # print("数据字典为：{}".format(tv_info))
    # print("数据字典长度为：{}".format(len(tv_info)))
    return tv_info


def change(server, server_dict, tv_na_dict, tv_ori_dict):
    tv_list = server_dict[server]
    # q.put(tv_list)
    filepath = os.getcwd() + r'/log.txt'
    all_num = len(tv_list)
    a = Cms(server)
    # q.put('登录cms：' + server)
    a.login()
    for ch in tv_list:
        print('completed: {}/{}'.format(tv_list.index(ch)+1, all_num))
        # q.put(ch)
        try:
            # q.put('开始修改'+ch)
            a.select_ch(ch)
            status = a.updata_ch(tv_name=tv_na_dict[ch], url=None, origin=tv_ori_dict[ch])
            # status = None
            # q.put(status)
            if status is None:
                print(ch+' 在服务器'+server+'是否更新成功: 成功')
                a.select_ch(ch)
                # with open(filepath, 'a') as f:
                #     f.write(datetime.datetime.today() + '      在服务器' + server + '更新成功: ' + ch + '\n')

            else:

                print(ch+' 在服务器'+server+'更新失败 ')
                with open(filepath, 'a') as f:
                    f.write(str(datetime.datetime.today()) + '      在服务器' + server + '更新失败: ' + ch + '\n')

        except (Exception, BaseException) as e:
            print('%s 更新 %s 信息失败' % (server, ch))
            print('#########################################################')
            print('str(Exception):\t', str(Exception))
            print('str(e):\t\t', str(e))
            print('repr(e):\t', repr(e))
            print('e.message:\t', e.message)
            print('traceback.print_exc():', traceback.print_exc())
            print('traceback.format_exc():\n%s' % traceback.format_exc())
            print('##########################################################')
            with open(filepath, 'a') as f:
                f.write(datetime.datetime.today() + '      在服务器' + server + '更新失败: ' + ch + '\n')
            continue


if __name__ == '__main__':
    # q = multiprocessing.Manager().Queue()
    wb = openpyxl.Workbook()
    sh = wb.active
    filepath = os.getcwd() + r'/log.txt'
    server_dict = {}
    server_list = []
    tv_na_dict = {}
    tv_ori_dict = {}
    tv_ser_dict = {}
    # print(server_list)
    Tv_info = get_ch()
    ch_list = list(Tv_info.keys())

    b = MyDb()
    conn = b.connect_db()
    tv_name = None
    origin = None
    server = None
    cpcode_list = []
    cpcode_dict = defaultdict(list)
    print('ch_list:{}'.format(ch_list))
    for TV_ch in ch_list:
        c = Tv_info[TV_ch]
        tv_list = []
        # print(type(c))
        if c is not None:
            ser_ip = re.findall(r'(?<=ser_ip=)[^,]+', c)
            if len(ser_ip) != 0:
                # Specified IP
                ip = ser_ip[0]
                # server_dict:{'54.78.82.194': ['TV14329']}
                tv_list, server_dict = b.get_info(TV_ch)
                cpcode_list.extend(tv_list)
                if ip in server_dict.keys():
                    cpcode_dict[ip].extend(server_dict[ip])

            else:
                tv_list, server_dict = b.get_info(TV_ch)
                server_list = server_dict.keys()
                cpcode_list.extend(tv_list)
                for ip in server_dict.keys():
                    if ip not in cpcode_dict.keys():
                        cpcode_dict[ip] = server_dict[ip]
                    else:
                        cpcode_dict[ip].extend(server_dict[ip])

            tv_n = re.findall(r'(?<=tv_name=)[^,]+', c)
            if len(tv_n) != 0:
                tv_name = tv_n[0]
            for t in tv_list:
                tv_na_dict[t] = tv_name

            ori = re.findall(r'(?<=origin=)[^,]+', c)
            if len(ori) != 0:
                origin = ori[0]
            for t in tv_list:
                tv_ori_dict[t] = origin

    # print("server_list：{}".format(server_list))
    # print("server_dict：{}".format(server_dict))
    print("tv_na_dict：{}".format(tv_na_dict))
    print("tv_ori_dict：{}".format(tv_ori_dict))
    b.close(conn)

    pool = multiprocessing.Pool(3)
    # print(multiprocessing.cpu_count()-1)

    print('开始多进程修改')
    results = []
    print('cpcode_dict:{}'.format(cpcode_dict))
    # print('server_list:{}'.format(server_list))
    for server in cpcode_dict.keys():
        # if server == '185.79.153.34' or server == '185.79.153.44':
        #     pass
        # else:
            print('开始修改{}上的节目\n'.format(server))
            # change(server, server_dict, tv_na_dict, tv_ori_dict)
            c = pool.apply_async(change, args=(server, cpcode_dict, tv_na_dict, tv_ori_dict))
            c.get()
    pool.close()
    pool.join()
    # print('q.qsize:')
    # print(q.qsize())
    # while not q.empty():
    #     print(q.get())
    # print('\n多进程修改完毕')
    # print('\n中继153.34和153.44请手动修改')


