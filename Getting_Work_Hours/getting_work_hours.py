# !/usr/bin/python3
# -*- coding: utf-8 -*-
'''
@File    :  getting work hours.py  
@Software:  PyCharm
@Modify Time:   2019/8/6 9:31
@Author:    LGH
@Version    0.1
@Desciption: 从表格中获取起始时间计算时长并写入新表
------------      -------    --------    -----------
'''

import openpyxl
import datetime
import os


def getting_data():
    path = os.getcwd()
    # print(path)
    file_path = path + '/打卡记录.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    # print(wb.sheetnames)
    # str = ws.cell(3, 1).value
    # print(str)
    # print(str[:5])
    # print(str[-6:-1])
    # print(len(str[-6:-1]))
    # print(str.split('\n'))

    all_person = []
    cols = ws.max_column
    rows = ws.max_row

    for num in range(2, rows, 2):
        person = {}
        if num + 1 > rows:
            break
        # print('num:' + str(num))
        job_num = ws.cell(num, 3).value
        # print(job_num)
        person['job_num'] = job_num
        name = ws.cell(num, 11).value
        # print(name)
        person['name'] = name
        work_time = {}
        for col in range(1, cols):
            # print('col:' + str(col))
            work_time1 = ws.cell(num+1, col).value
            if work_time1 is not None:
                star_time_str = work_time1[:5]
                # print(star_time_str)
                end_time_str = work_time1[-6:-1]
                # print(end_time_str)
                star_time = datetime.datetime.strptime(star_time_str, "%H:%M")
                # print(star_time)
                end_time = datetime.datetime.strptime(end_time_str.strip(r'\n'), "%H:%M")
                # print(end_time)
                work_times = str(end_time - star_time)
                # print(work_times)
                work_time[col] = work_times
            else:
                # print(work_time1)
                work_time[col] = work_time1
        person['work_time'] = work_time
        all_person.append(person)
    # print('num:' + num)
    return all_person


def create_excle(persons):
    path = os.getcwd()
    Path = path + '/results.xlsx'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    # 写入第一行
    cols = len(persons[0]['work_time'])
    for col in range(cols):
        # print(col)
        ws1.cell(1, col+1, col+1)
    #每个person 占两行
    excle_num = len(persons)*2
    print(excle_num)
    num = 0
    for row1 in range(2, excle_num+2, 2):
        ws1.cell(row1, 1, '工号:')
        ws1.cell(row1, 2, persons[num]['job_num'])
        ws1.cell(row1, 3, '姓名：')
        ws1.cell(row1, 4, persons[num]['name'])
        # print(row1)
        # print(persons[num]['name'])
        for col in range(cols):
            ws1.cell(row1+1, col+1, persons[num]['work_time'][col+1])
        num = num+1

    wb.save(Path)


if __name__ == "__main__":
    all_persons = getting_data()
    # print(all_persons)
    create_excle(all_persons)