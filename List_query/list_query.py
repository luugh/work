#!/usr/bin/python
# -*- coding:utf-8 -*-
# by lgh
#

import requests
import openpyxl
import os
import json
import time


class Cms:
    def __init__(self):
        self.session = None

    def cms_login(self):
        url = 'http://www.haishilist.com:8180/cms/login.action'
        path = os.getcwd() + r'\login.txt'
        line_list = []
        try:
            with open(path) as f:
                lines = f.readlines()
                for line in lines:
                    if line == '\n':
                        pass
                    else:
                        line_list.append(line)
            # print(line_list)
            name = line_list[0].strip()
            password = line_list[1].strip().replace("'", "")
            # print(name, password)
        except Exception as err:
            print(err)
            print('当前目录下未找到文件：login.txt')

        data = {
            'request_locale': 'en_US',
            'userName': name,
            'userPassword': password,
            'checkbox': ''
        }
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0',
            'Referer': 'http://www.haishilist.com:8180/cms/login.action',
            'Connection': 'keep-alive'
        }
        self.session = requests.Session()
        res = self.session.post(url=url, headers=headers, data=data)
        # print(res.status_code)

    def select_ch(self, ch_list):
        ch_dict = {}
        url = 'http://www.haishilist.com:8180/cms//channel/channelList.action?hasInput=N'
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0',
            'Referer': 'http://www.haishilist.com:8180/cms/login.action',
            'Connection': 'keep-alive'
        }
        for ch in ch_list:
            # time.sleep(1)
            data = {
                'page': '1',
                'rp': '1000',
                'sortname': 'channelnumber',
                'sortorder': 'desc',
                'query': '',
                'qtype': '',
                'servicbean.queryWords': ch,
                # 'servicbean.queryWords': 'TV77051',
                'servicbean.type': '-1',
                'servicbean.status': '-1',
                'status': '0',
                'releaseStatus': '0',
                'broadcasttype': '-1',
                'ids': ''
            }
            res = self.session.post(url=url, headers=headers, data=data)
            # print(res.text)
            # print(res.status_code)
            res_dict = json.loads(res.text)
            # print(res_dict)
            # print(type(res_dict['rows']))
            if len(res_dict['rows']) == 0:
                print('%s 不存在于列表中' % ch)
                ch_dict[ch] = '不在列表中'
            else:
                # print('%s 在列表中' % ch)
                keywords = self.judge_status(res_dict)
                print(ch, keywords)
                ch_dict[ch] = keywords

        return ch_dict

    @classmethod
    def judge_status(cls, res_dict):
        num = res_dict['total']
        # print(num)
        # print(res_dict)
        rows_list = res_dict['rows']
        # print('len(rows_list):{}'.format(len(rows_list)))
        status = rows_list[0]['releaseStatus']
        values = '0'
        if num == 1:
            values = '1'
        else:
            for i in range(1, num):
                # print('i:{}'.format(i))
                if status == rows_list[i]['releaseStatus']:
                    values = '1'
                else:
                    values = '2'
                    break
        if values == '1':
            if status == '1':
                keywords = '列表已上线'
            else:
                keywords = '列表已下线'
        else:
            keywords = '列表下发状态不一致'
        return keywords


def create_excl(ch_dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ch_list = list(ch_dict.keys())
    # print(ch_list)
    print('结果写入表格：result.xlsx')
    for num in range(0, len(ch_list)):
        ws.cell(num+1, 1, ch_list[num])
        ws.cell(num+1, 2, ch_dict[ch_list[num]])
    file_path = os.getcwd() + r'\result.xlsx'
    wb.save(file_path)


def get_ch():
    file_path = os.getcwd() + '\initial.xlsx'
    print(file_path)
    wb = openpyxl.load_workbook('.\initial.xlsx')
    ws = wb.active
    ch_list = []
    rows = ws.max_row
    # print(rows)
    for row in range(1, rows+1):
        ch_list.append(ws.cell(row=row, column=2).value)
    # print(ch_list)
    return ch_list


if __name__ == '__main__':
    list1 = get_ch()
    a = Cms()
    a.cms_login()
    re_dict = a.select_ch(list1)
    create_excl(re_dict)
