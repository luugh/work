# !/usr/bin/python3
# -*- coding: utf-8 -*-

# @Time    : 2021/3/15 9:35
# @Author  : LGH
# @File    : africanews.py
# @Software: PyCharm


from selenium import webdriver
from selenium.webdriver.firefox.options import Options
import requests
import json
# from africanwes.SE import Cms
from SE import Cms
import openpyxl
import os
import time


class africanews:
    def __init__(self):
        self.url = 'https://www.africanews.com/live/'
        self.cookies = ''

    def get_cookies(self):

        # chrome_options = Options()
        # chrome_options.add_argument('--headless')
        # # 让Chrome在root权限下启动
        # chrome_options.add_argument('--no-sandbox')
        # chrome_options.add_argument("window-size=1024,768")
        # chrome_options.add_argument('--disabel-dev-shm-usage')
        # #
        # browser = webdriver.Chrome(options=chrome_options)
        # headers = {
        #     "Host": "www.africanews.com",
        #     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:87.0) Gecko/20100101 Firefox/87.0",
        #     "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        #     "Accept-Language": "zh-CN",
        #     "Accept-Encoding": "gzip, deflate, br",
        #     "Connection": "keep-alive",
        #     "Upgrade-Insecure-Requests": "1",
        #     "Cache-Control": "max-age=0",
        #     "TE": "Trailers"
        # }
        # browser = webdriver.Chrome()
        # browser.get(url=self.url, headers=headers)

        firefox_options = Options()
        firefox_options.add_argument("--headless")

        browser = webdriver.Firefox(firefox_options=firefox_options)
        try:
            browser.get(url=self.url)
            # 在登录时只使用的到了'name'和'value'
            # AMCV_237E3E1059F2E95E0A495E5D@AdobeOrg: -408604571|MCIDTS|18702|vVersion|4.6.0
            print('>>>browser.get_cookies:')
            print(browser.get_cookies())
            # print(browser.title)
            self.cookies = browser.get_cookies()[0]
        except Exception as e:
            print(e)
        finally:
            browser.close()

    def get_cookie(self):
        url = 'https://www.africanews.com/live/'
        headers = {
            "Host": "www.africanews.com",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:87.0) Gecko/20100101 Firefox/87.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "zh-CN",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Cache-Control": "max-age=0",
            "TE": "Trailers"
        }
        session = requests.Session()
        # 超时重试

        res = session.get(url=url, headers=headers)
        print('session.cookies >>> {}'.format(session.cookies))

    def get_palylink(self):
        session = requests.session()
        cookies_json = json.dumps(self.cookies)
        print(f">>>>>>>>>>>>>cookies_json:{cookies_json}")
        headers = {
            "Host": "www.africanews.com",
            "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv":"87.0) Gecko/20100101 Firefox/87.0',
            "Accept": "*/*",
            "Accept-Language": "zh-CN",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
            "Referer": "https://www.africanews.com/live/",
            # "Cookie": " AMCV_237E3E1059F2E95E0A495E5D%40AdobeOrg=-408604571%7CMCIDTS%7C18702%7CvVersion%7C4.6.0",
            # "Cookie": "{}:{}".format(self.cookies['name'], self.cookies['value']),
            "Cookie": cookies_json,
            "Cache-Control": "max-age=0",
            "TE": "Trailers"
        }
        # print('headers.Cookies:{}'.format(headers['Cookie']))
        url2 = 'https://www.africanews.com/api/watchlive.json'
        res1 = session.get(url=url2, headers=headers)
        # print(res.text)
        url_dict = json.loads(res1.text)
        # print(url_dict)
        print('>>>获取 hls_link:')
        hls_link = "http:{}".format(url_dict['url'])

        res2 = session.get(hls_link)
        res2_m3u8 = json.loads(res2.text)
        print('>>>获取返回的m3u8:')
        print(res2_m3u8)
        if len(res2_m3u8['primary']) > 10:
            play_m3u8 = res2_m3u8['primary']
        elif len(res2_m3u8['backup'])> 10:
            play_m3u8 = res2_m3u8['backup']
        else:
            print('获取m3u8文件失败！')

        res3 = session.get(play_m3u8)
        # print(res3.text)
        # print(type(res3.text))
        res3_list = res3.text.split()
        print('>>>获取播放列表 res3_list')
        print(res3_list)
        for i in res3_list:
            if '720p20' in i:
                end_l = i
        play_link = play_m3u8.replace('playlist.m3u8', end_l)
        print(play_link)
        return play_link


def modify_excel(file_path, link):
    w1 = openpyxl.load_workbook(file_path)
    s1 = w1.active
    s1.cell(3, 3, link)
    cpcode = s1.cell(3,1).value
    w1.save(file_path)

    return cpcode


if __name__ == '__main__':
    print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
    time1 = time.time()
    print(time.ctime())
    path = os.getcwd()
    file_path = path + r'/africanews.xlsx'
    a = africanews()
    print('>>>geting cookies')
    # a.get_cookie()
    a.get_cookies()
    print('>>>geting playlink')
    play_link = a.get_palylink()
    print('>>>modify excel')
    cp = modify_excel(file_path, play_link)

    b = Cms('xxx.xxx.xxx.xxx')
    b.login()
    print('>>>up file to SE')
    b.up_tr_file(file_path)
    print('>>>select ch status')
    res_text = b.select_tr_ch(cp)

    Id = json.loads(res_text)['list'][0]['id']
    b.issue(Id)
    n = 0
    while True:
        res1_text = b.select_tr_ch(cp)
        issue_state = json.loads(res1_text)['list'][0]['issue_state']
        print('>>>获取下发状态(未下发:105;已下发:106):')
        print('issue_state:{}'.format(issue_state))
        if issue_state == 105:
            n += 1
            print('issue failed,retry {}'.format(n))
            res1 = b.select_tr_ch(cp)
            Id = json.loads(res_text)['list'][0]['id']
            b.issue(Id)
            time.sleep(5)
        elif issue_state == 106:
            print('下发成功!退出程序！')
            break
    time2 = time.time()
    print(time2 - time1)
