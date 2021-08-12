# !/usr/bin/python3
# -*- coding: utf-8 -*-

# @Time    : 2020/11/5 14:09
# @Author  : LGH
# @File    : auto_update.py
# @Software: PyCharm

import openpyxl
import os
import requests
import datetime
import json
import time


class Cms:
    def __init__(self, ip):
        self.ip = ip
        self.session = ''
        self.rows = ''

    def login(self):
        ip = self.ip
        url = 'http://%s:8088/user/login' % ip
        host = '%s:8088' % ip
        referer = 'http://%s:8088/user/login' % ip
        headers = {
            # 'Host': '178.132.6.62:8088',
            'Host': host,
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/78.0.3904.108 Safari/537.36',
            'Accept': "application/json",
            'Accept-Language': 'zh,en-US;q=0.7,en;q=0.3',
            'Accept-Encoding': 'gzip, deflate',
            'Content-Type': 'application/json; charset=utf-8',
            'Referer': referer,
            'Connection': 'keep - alive',
            # 'DNT': '1',
            'origin': 'http://%s:8088' % ip,
            # 'Cookie': self.cookie
        }
        login_url = 'http://%s:8088/cdn/submitLogin' % ip
        filepath = os.getcwd() + r'/login.txt'
        with open(filepath, 'r') as f:
            list1 = f.readlines()
        name = list1[0].strip()
        passwd = list1[1].strip()
        # name = 'admin'
        # passwd = 'Q5t0wRLN'
        login_post = {

            "userName": name,
            "password": passwd,
            "type": "account",
            "local": "en_US",
        }
        # print(login_post)
        self.session = requests.Session()
        # 注意此时用的json参数
        res = self.session.post(url=login_url, json=login_post, headers=headers)

        # SHIROSESSION = 97229d9a-912b-4a19-a685-9cef7a3fddd4;Path = /;
        # print(res.headers['Set-Cookie'])
        cookie = res.headers['Set-Cookie']
        # cookie = 'SHIROSESSION=97229d9a-912b-4a19-a685-9cef7a3fddd4;Path = /'
        # 正向后视断定
        # test = re.findall(r'(?<=SHIROSESSION=)[^;]+', cookie)

        # print(test)
        # print(self.cookie)
        # print(res.status_code)
        # print(res.text)
        # print(res.headers)
        # print(res.headers['date'])

    def sync_ch(self):
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0'
        }
        url = 'http://{}:8088/cdn/record/sync'.format(self.ip)
        self.session.get(url)
        time.sleep(18)

    # get channels status in batch
    def select_batch(self, ip):
        # dict_origin1 = {
        #     '185.79.153.34': '第三方',
        #     '199.167.137.146': '青蛙源',
        #     '185.79.153.44': '荷兰第三方',
        #     '154.54.220.98': '转码源',
        #     '54.78.82.194': 'bestTV',
        #     '54.78.82.194': '备用',
        #     '38.94.97.150': '第三方'
        # }
        dict_origin = {
            '第三方': '185.79.153.34',
            '青蛙源': '199.167.137.146',
            '荷兰第三方': '185.79.153.44',
            '转码源': '154.54.220.98',
            'bestTV': '54.78.82.194',
            '备用': '54.78.82.194',
            '第三方_': '38.94.97.150'
        }
        print('ip: {}'.format(ip))
        ori = [x for x in dict_origin.keys() if dict_origin[x] == ip][0]
        print('ori: {}'.format(ori))
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0'
        }

        # 'http://54.78.82.194:8088/cdn/record?seconds=300&state=101&origin=bestTV&current=2&pageSize=1000'
        url = 'http://{}:8088/cdn/record?seconds=300&state=101&origin={}&current=1&pageSize=1000'.format(ip, ori)
        res = self.session.get(url=url, headers=headers)
        # print(res.text)
        res_dict = json.loads(res.text)
        ch_state = {}
        # 判断是否有结果
        # print(type(res_dict['pagination']['total']))
        if res_dict['pagination']['total'] > 0:
            for ch_dict in res_dict['list']:
                state = ch_dict['state']
                if state == 101:
                    ch_state[ch_dict['channel_id']] = '中断'
                else:
                    ch_state[ch_dict['channel_id']] = state
        print('ch_state:{}'.format(ch_state))
        return ch_state

    def select_interrupt(self):
        dict2 = {
            '印度': '185.79.153.34',
            '青蛙源_南美': '199.167.137.146',
            '葡语西语44': '185.79.153.44',
            'myHD': '154.54.220.98',
            'Telefoot': '185.79.153.34',
            'bestTV': '54.78.82.194',
            'bestTV备用': '54.78.82.194',
            'xcstream': '38.94.97.150'
        }
        headers = {
                'Host': '38.94.97.150:8088',
                'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0',
                'Accept': '*/*',
                'Accept-Language': 'zh-CN',
                'Accept-Encoding': 'gzip, deflate',
                'Referer': 'http://38.94.97.150:8088/live/recording',
                'Connection': 'keep-alive',
                # 'Cookie': 'WEBJSESSIONID=2b58405d-728f-430d-b589-5f35bd337e35'

        }

    # def select_ch(self, cpcode):
    #     headers = {
    #         'Host': '%s:8088' % self.ip,
    #         'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0',
    #         'Accept': 'application/json, text/javascript, */*; q=0.01',
    #         'Accept-Language': 'zh,en-US;q=0.7,en;q=0.3',
    #         'Accept-Encoding': 'gzip, deflate',
    #         'Referer': 'http://%s:8088/live/recording' % self.ip,
    #         # 'Content-Type': 'application/json; charset=utf-8',
    #         # 'X-Requested-With': 'XMLHttpRequest',
    #         # 'Content-Length': '113',
    #         # 'DNT': '1',
    #         'Connection': 'keep-alive',
    #         # 'Cookie': self.session
    #     }
    #     select_url = 'http://' + self.ip + ':8088/cdn/record?channel_id=' + cpcode + '&pageSize=' + '100'
    #     res = self.session.get(url=select_url, headers=headers)
    #     result = json.loads(res.text)
    #     print('result: {}'.format(result))
    #     if len(result['list']) != 0:
    #         cpcode = result['list'][0]['channel_id']
    #         state = result['list'][0]['state']
    #         result1 = result['list'][0]
    #         self.rows = result1
    #         # print('唯一标识：' + self.rows['channel_id'] + '\n' + '节目名称：' + self.rows['channel_name'] +
    #         #       '\n' + '来源分组：' + self.rows['origin'] + '\n' + 'state: ' + str(self.rows['state']))
    #         s_state = ''
    #         if state == 101:
    #             s_state = '中断'
    #         elif state == 100:
    #             s_state = ''
    #         else:
    #             s_state == state
    #         return s_state
    #     else:
    #         print('list为空，未查询到信息')
    #         return None


# get status for channels
def get_ch(sh_name, wb, dict2):
    today = datetime.datetime.today()
    today_str = today.strftime('%Y/%m/%d')
    print(today_str)
    sh = wb[sh_name]
    rows = sh.max_row
    cols = sh.max_column
    print("获取到{}行数据".format(rows))
    print("获取到{}列数据".format(cols))
    list_title = [sh.cell(row=1, column=col+1).value for col in range(cols)]
    # print(sh.cell(1, 1).value)
    # print(list_title)
    # 去除list_title尾部的None元素
    while list_title[cols-1] is None:
        list_title.pop()
        cols = len(list_title)
    # print('去除None元素后list_title: {}'.format(list_title))
    # 定位'唯一标示'所在列序号
    col1 = list_title.index('唯一标示')+1
    # print(col1)
    # 查看当前记录的最近的日期并更新日期
    date1 = list_title[cols-1]
    n = 1
    print('date1: {}'.format(date1))
    print('today_str: {}'.format(today_str))
    # print(list_title[10].strftime('%Y/%m/%d'))
    print('list_title: {}'.format(list_title))
    last_day = list_title[-1].strftime('%Y/%m/%d')
    print('last_day: {}'.format(last_day))
    if last_day != today_str:
        # 更新日期
        while date1.strftime('%Y/%m/%d') < today_str:
            # print(date1)
            date1 += datetime.timedelta(days=n)
            list_title.append(date1)
        # 获取唯一标识以及节目状态
        ch_list = []
        for row in range(rows):
            if row + 2 < rows:
                ch_list.append(sh.cell(row+1, col1).value)
        # print('ch_list: {}'.format(ch_list))
        ip = dict2[sh_name]
        a = Cms(ip)
        a.login()
        a.sync_ch()
        ch_state = a.select_batch(ip)
        ch_in = ch_state.keys()
        state_list = []
        for ch in ch_list:
            if ch is not None:
                # pass
                if ch in ch_in:
                    state_list.append(ch_state[ch])
                else:
                    state_list.append(None)
                # pass
            else:
                state_list.append(None)
        print('state_list:'.format(state_list))
        return state_list, list_title
    else:
        state_list = {}
        list_title = {}
        return state_list, list_title


def write_to_excel(sh_name, wb, state_list, list_title):
    sh = wb[sh_name]
    cols = len(list_title)
    for num in range(cols):
        sh.cell(1, num+1, list_title[num])
    rows = len(state_list)
    for num in range(1, rows):
        if state_list[num] is not None:
            sh.cell(num+1, cols, state_list[num])

    return wb


if __name__ == "__main__":
    file_list = os.listdir()
    for file in file_list:
        if '第三方源' in file:
            file_name = file.strip()
            print(file_name)
    file_path = os.getcwd() + '/' +file_name
    print("表格路径为：{}".format(file_path))
    wb = openpyxl.load_workbook(file_path)
    sh_names = wb.sheetnames
    print('#############################')
    print('sh_names: {}'.format(sh_names))
    # ['印度', '青蛙源_南美', '葡语西语44', 'myHD-shahidTV', '英国客户TECHNOMATE', 'Romania_hamid']
    dict1 = {
        '印度': '185.79.153.34',
        '青蛙源_南美': '199.167.137.146',
        '葡语西语44': '185.79.153.44',
        'myHD': '154.54.220.98',
        'Telefoot': '185.79.153.34',
        # 'bestTV': '54.78.82.194',
        # 'bestTV备用': '154.54.220.98',
        'xcstream': '38.94.97.150'
    }
    # for sh_name in sh_names:
    #     get_ch(sh_name, wb, dict1)
    for sh_name in sh_names:
        print('sh_name: {}'.format(sh_name))
        state_list, list_title = get_ch(sh_name, wb, dict1)
        if len(state_list) != 0:
            wb = write_to_excel(sh_name, wb, state_list, list_title)

    wb.save(file_path)
    print('表格写入完成')
