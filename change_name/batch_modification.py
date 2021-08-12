#!/usr/bin/python3
# -*- coding:utf-8 -*-
# by lgh
# change channel's name


from cms_login import Cms
import pymysql
import os
import openpyxl
import re
import traceback


class MyDb:
    def __init__(self):
        self.name = 'channelaction'
        self.passwd = 'd0177951d1493f49'
        self.dbname = 'cmdb'
        self.port = '3306'
        self.host = '208.81.204.10'

    def connect_db(self):
        db = pymysql.connect(self.host, self.name, self.passwd, self.dbname)
        return db

    def get_info(self, connect_db, name):
        print(name)
        cursor = connect_db.cursor()
        select_command = u"select hostIp from cdnChannel where channelMark='%s';" % name
        cursor.execute(select_command)
        results = cursor.fetchall()
        # print(type(results))
        # print(results)
        ip_list = []
        for i in results:
            ip_list.append(i[0])
        print(ip_list)
        # 返回ip列表元组
        return ip_list

    def close(self, connect_db):
        connect_db.close()


def get_ch():
    tv_info = {}
    file_path = os.getcwd()+'\Modify_information.xlsx'
    print(file_path)
    wb = openpyxl.load_workbook(file_path)
    sh = wb.active
    rows = sh.max_row
    print(rows)
    for row in range(rows):
        tv_info[sh.cell(row=row+1, column=1).value] = sh.cell(row=row+1, column=2).value
    print(tv_info)
    print(len(tv_info))
    return tv_info


if __name__ == '__main__':
    wb = openpyxl.Workbook()
    sh = wb.active
    filepath = os.getcwd() + '\log.txt'

    Tv_info = get_ch()
    ch_list = list(Tv_info.keys())
    # print(ch_list)
    for TV_name in ch_list:

        b = MyDb()
        conn = b.connect_db()
        # TV_name = 'TV1508'
        Ip_list = b.get_info(conn, TV_name)
        b.close(conn)
        tv_name = None
        origin = None
        # b = 'tv_name=asfdaer,origin=adsfadfwer,url=fadfwerfsf'
        # b = 'origin=adsfadfwer,url=fadfwerfsf'
        # print(re.findall(r'tv_name=[^,]', Tv_info[TV_name]))
        # print(re.findall(r'(?<=tv_name=)[^,]+', b))
        c = Tv_info[TV_name]
        tv_n = re.findall(r'(?<=tv_name=)[^,]+', c)
        # print(len(tv_n))
        if len(tv_n) != 0:
            tv_name = tv_n[0]
            with open(filepath, 'a') as f:
                f.write(TV_name+' 更新名称 ')
            # print(tv_name)
        # print(re.findall(r'(?<=origin=)[^,]+]', b))
        # print(len(re.findall(r'(?<=origin=)[^,]+]', b)))
        ori = re.findall(r'(?<=origin=)[^,]+', c)
        if len(ori) != 0:
            origin = ori[0]
            with open(filepath, 'a') as f:
                f.write(TV_name+' 更新来源分组')
            # print(origin)
        if len(Ip_list) != 0:
            for ip in Ip_list:
                try:
                    a = Cms(ip)
                    a.login()
                    a.select_ch(TV_name)
                    # print(Tv_info[TV_name])
                    status = a.updata_ch(tv_name=tv_name, url=None, origin=origin)
                    # status = 'true'
                    with open(filepath, 'a') as f:
                        f.write('      在服务器'+ip+'是否更新成功: '+status+'\n')
                    a.select_ch(TV_name)
                except (Exception, BaseException) as e:
                    print('%s 更新 %s 信息失败' % (ip, TV_name))
                    print('#########################################################')
                    print('str(Exception):\t', str(Exception))
                    print('str(e):\t\t', str(e))
                    print('repr(e):\t', repr(e))
                    # print('e.message:\t', e.message)
                    # print('traceback.print_exc():', traceback.print_exc())
                    # print('traceback.format_exc():\n%s' % traceback.format_exc())
                    print('##########################################################')
                    continue

        else:
            with open(filepath, 'a') as f:
                f.write(TV_name+'节目服务器未添加\n')
            print('%s 节目服务器未添加' % TV_name)
