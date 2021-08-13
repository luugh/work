#!/usr/bin/python3
# -*- coding: utf-8 -*-

# @Time    : 2021/2/6 12:06
# @Author  : LGH
# @File    : change_name.py
# @Software: PyCharm

import asyncio
import json

from SE import Cms
from MyDb import MyDb
import openpyxl
import pymysql
from collections import defaultdict
import re
import time

def get_ch():
    w1 = openpyxl.load_workbook('new.xlsx')
    sh1 = w1.active
    rows = sh1.max_row
    # print(rows)
    # print(type(rows))
    ch_list = []
    for n in range(1, rows+1):
        if sh1.cell(n+1, 1).value is not None and 'TV' in sh1.cell(n+1, 1).value:
            ch_info = {}
            ch_info.setdefault('ch_code', sh1.cell(n+1, 1).value)
            ch_info.setdefault('ch_name', sh1.cell(n+1, 2).value)
            ch_info.setdefault('ch_ori', sh1.cell(n+1, 3).value)
            ch_info.setdefault('ch_ip', sh1.cell(n+1, 4).value)
            ch_info.setdefault('ch_op', sh1.cell(n+1, 5).value)
            ch_info.setdefault('ch_link', sh1.cell(n+1, 6).value)
            ch_list.append(ch_info)
    return ch_list


async def modefy_task(ip, cpcode_dict, real_cp):
    b=Cms(ip)
    await b.login()
    for cp in cpcode_dict[ip]:
        res1 = await b.select_ch(cp)
        # print(type(res1))
        if type(res1).__name__ != 'dict':
            print(res1)
            continue
        # print(f'time: {time.time()}  res: {res1}')
        if real_cp[cp]['ch_name'] is None:
            channel_name = res1['ch_name']
        else:
            channel_name = real_cp[cp]['ch_name']
        if real_cp[cp]['ch_ori'] is None:
            origin = res1['origin']
        else:
            origin = real_cp[cp]['ch_ori']
        data = {
            "channel_id": cp,
            "channel_name": channel_name,
            "id": res1['id'],
            "issue_state": res1['issue_state'],
            "on_demand": res1['on_demand'],
            "origin": origin,
            "select_app": res1['select_app'],
            "source_url": res1['source_url'],
            "store_time": res1['store_time']
        }
        # print(f'data: {data}')
        res2 = await b.modify_ch(data)
        res2 = json.loads(res2)
        # print(type(res2['code']))
        # print(f'modefy result: {res2}')
        if res2['code'] == 200:
            res3 = await b.select_ch(cp)
            print('>>>>>>>>>>>>>>>>>>')
            print(f'修改前信息: {res1}')
            print(f'修改后信息: {res3}')
            print(f'在服务器 {ip} 上节目 {cp} 信息修改成功')
            print('>>>>>>>>>>>>>>>>>>')
        else:
            print('<<<<<<<<<<<<<<<<<<<<')
            print(f'在服务器 {ip} 上节目 {cp} 信息修改失败，报错信息：{res2}')
            print('<<<<<<<<<<<<<<<<<<<<')

async def modefy(cpcode_dict, real_cp):
    tasks = []
    for ser in cpcode_dict.keys():
        task = asyncio.create_task(modefy_task(ser, cpcode_dict, real_cp))
        tasks.append(task)
    await asyncio.gather(*tasks)


if __name__ == '__main__':
    ch_list = get_ch()
    print(f'ch_list: {ch_list}')
    cpcode_dict = defaultdict(list)
    cpcode_list = []
    a = MyDb()
    conn = a.connect_db()
    time1 = time.time()
    real_cp = {}
    for ch in ch_list:
        if ch['ch_ip'] is not None and len(ch['ch_ip']) > 6:
            name_l, server_d = a.get_info(ch['ch_code'])
            if ch['ch_ip'] in server_d.keys():
                cpcode_dict[ch['ch_ip']].extend(server_d[ch['ch_ip']])
        else:
            name_l, server_d = a.get_info(ch['ch_code'])
            for ip in server_d.keys():
                cpcode_dict[ip].extend(server_d[ip])
        cp_real = {}
        for name in name_l:

            cp_real.setdefault('ch_name', ch['ch_name'])
            cp_real.setdefault('ch_ori', ch['ch_ori'])
            cp_real.setdefault('ch_ip', ch['ch_ip'])
            cp_real.setdefault('ch_op', ch['ch_op'])
            cp_real.setdefault('ch_link', ch['ch_link'])
            real_cp.setdefault(name, cp_real)
    time2 = time.time()
    # print(f'spend time: {time2-time1}')
    # print(f'name_l: {name_l}')
    # print(f'server_d: {server_d}')
    # print(f'cpcode_dict: {cpcode_dict}')
    # print(f'real_cp: {real_cp}')
    a.close(conn)
    start = time.time()
    asyncio.run(modefy(cpcode_dict, real_cp))
    end = time.time()
    print(f'spend: {end-start}')
    