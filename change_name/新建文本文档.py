#! /usr/bin/python3
# -*-coding:utf-8-*-
# by lgh
# @Time     :2019/3/20 9:07
# @Author   :LGH
# @Site     :
# @File     :batch_modification0.1.py
# @Software :

import requests
import json
import time
import pymysql
import os
import openpyxl
import re
import traceback
import multiprocessing


class MyDb:
    def __init__(self):
        self.name = 'channelaction'
        self.passwd = 'd0177951d1493f49'
        self.dbname = 'cmdb'
        self.port = '3306'
        self.host = '208.81.204.10'

    def connect_db(self):
        db = pymysql.connect(self.host, self.name, self.passwd, self.dbname)
        return db


    def get_info(self, connect_db, name):
        print(name)
        cursor = connect_db.cursor()
        select_command = u"select hostIp from cdnChannel where channelMark='%s';" % name
        cursor.execute(select_command)
        results = cursor.fetchall()
        # print(type(results))
        # print(results)
        ip_list = []
        for i in results:
            ip_list.append(i[0])
        #print(ip_list)
        # 返回ip列表元组
        return ip_list

    def close(self, connect_db):
        connect_db.close()


class Cms:

    def __init__(self, ip):
        self.ip = ip
        self.cookie = ''
        self.rows = {}

    def login(self):
        # get cookies
        ip = self.ip
        url = 'http://%s:8180/cms/login.action' % ip
        # print(url)
        re = requests.get(url)
        # print(re.headers)
        # print(type(re.headers['Set-Cookie']))
        jsession_id = re.headers['Set-Cookie']

        host = '%s:8081' % ip
        referer = 'http://%s:8180/cms/goLogin.action' % ip
        # cookie = jsession_id + ";cms_language=en_US; loginUser=%E5%88%98%E5%86%A0%E5%8D%8E; loginPass=zVFaXk37"
        self.cookie = jsession_id
        headers = {
            # 'Host': '178.132.6.62:8180',
            'Host': host,
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0',
            'Accept': "text/css,*/*;q=0.1",
            'Accept-Language': 'zh,en-US;q=0.7,en;q=0.3',
            'Accept-Encoding': 'gzip, deflate',
            # 'Referer': 'http://178.132.6.62:8180/cms/goLogin.action',
            'Referer': referer,
            'DNT': '1',
            'Connection': 'keep-alive',
            'Cookie': self.cookie
        }

        login_url = 'http://%s:8180/cms/login.action' % ip
        filepath = os.getcwd()+'\login.txt'
        with open(filepath, 'r') as f:
            list = f.readlines()
        name = list[0].strip()
        passwd = list[1].strip()
        # name = 'admin'
        # passwd = '5dDrxumL'
        login_post = {
            'request_local': 'en_US',
            'userName': name,
            'userPassword': passwd,
            'checkbox': ''
        }
        res = requests.post(url=login_url, data=login_post, headers=headers)

        # print(res.status_code)
        # print(res.text)
        # print(res.headers)
        # print(res.headers['date'])

    def select_ch(self, cpcode):

        headers = {
            'Host': '%s:8081' % self.ip,
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0',
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'Accept-Language': 'zh,en-US;q=0.7,en;q=0.3',
            'Accept-Encoding': 'gzip, deflate',
            'Referer': 'http://%s:8180/cms//servermonitor/channelrecordlist.action' % self.ip,
            'Content-Type': 'application/x-www-form-urlencoded',
            'X-Requested-With': 'XMLHttpRequest',
            # 'Content-Length': '113',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Cookie': self.cookie
        }

        select_post = {
            'page': '1',
            'rp': '15',
            'sortname': 'id',
            'sortorder': 'desc',
            'query': '',
            'qtype': '',
            'name': cpcode,
            'status': '',
            'activeStatus': '',
            'streamStatus': '',
            'OnDemandFlag': '',
            'mypage': '1'
        }
        select_url = 'http://%s:8180/cms//servermonitor/getChannelRecords.action' % self.ip
        res = requests.post(url=select_url, data=select_post, headers=headers)
        # print(res.text)
        # print(type(res.text))
        # res.text输出内容为字符串，需要转换为字典
        result = json.loads(res.text)
        id = result['rows'][0]['id']
        # 获取id
        post_data = {'id': id}
        select_url1 = 'http://%s:8180/cms//servermonitor/getChannelRecordById.action' % self.ip

        ress = requests.post(url=select_url1, data=post_data, headers=headers)
        result1 = json.loads(ress.text)
        self.rows = result1
        # print(self.rows)
        # 提取唯一标识
        cpcode = self.rows['cpcontentid']
        channel_name = self.rows['remark']
        if self.rows['streamStatus'] =='1':
            stream_status = '正常'
        else:
            stream_status = '中断'
        status = self.rows['statusStr']
        origin = self.rows['origin']
        source_url = self.rows['sourceurl']
        print(('频道名称：%s\n' % channel_name)+('频道唯一标识：%s\n' % cpcode)+('流状态：%s\n' % stream_status)
              +('下发状态：%s\n' % status)+('源链接%s\n' % source_url)+('正在使用的链接：%s\n' % self.rows['currenturl'])
              +('备注：%s\n' % origin)+("\n"))
        # print('频道名称：%s' % channel_name)
        # print('频道唯一标识：%s' % cpcode)
        # print('流状态：%s' % stream_status)
        # print('下发状态：%s' % status)
        # print('源链接%s' % source_url):
        # print('正在使用的链接：%s' % self.rows['currenturl'])
        # print('备注：%s' % origin)
        # print("\n")

    def updata_ch(self, tv_name=None, url=None, origin=None):
        headers = {
            'Host': '%s:8081' % self.ip,
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0',
            'Accept': '*/*',
            'Accept-Language': 'zh,en-US;q=0.7,en;q=0.3',
            'Accept-Encoding': 'gzip, deflate',
            'Referer': 'http://%s:8180/cms//servermonitor/channelrecordlist.action' % self.ip,
            'Content-Type': 'application/x-www-form-urlencoded',
            'X-Requested-With': 'XMLHttpRequest',
            # 'Content-Length': '113',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Cookie': self.cookie
        }
        if tv_name is not None:
            self.rows['remark'] = tv_name
            print('更新名称')
        if url is not None:
            self.rows['sourceurl'] = url
            print('更新源url')
        if origin is not None:
            self.rows['origin'] = origin
            print('更新来源分组')
        # print(self.rows)
        updata_post = {
            'OnDemandFlag': self.rows['onDemandFlag'],
            'activeStatus': self.rows['activeStatus'],
            'broadcasttype': self.rows['broadcasttype'],
            'channelIds': self.rows['channelId'],
            'status': self.rows['status'],
            'applicationMointorId': self.rows['applicationMointorId'],
            'sourceurl': self.rows['sourceurl'],
            'storageduration': '',
            'addtype': '0',
            'name': self.rows['cpcontentid'],
            'remark': self.rows['remark'],
            'origin': self.rows['origin'],  # 来源分组
            'id': self.rows['id']
        }
        # print('updata_post:' + str(updata_post))
        update_url = 'http://%s:8180/cms//servermonitor/modifyChannelRecord.action' % self.ip

        res = requests.post(url=update_url, data=updata_post, headers=headers)
        # print(res.text)
        # status = 'false'
        # if res.text == 'true':
        #     q.put(('------------------------------')+('%s上%s更新成功\n' % (self.ip, self.rows['cpcontentid']))
        #           +('------------------------------\n'))
        #     # print('%s上%s更新成功' % (self.ip, self.rows['cpcontentid']))
        #     # print('------------------------------')
        # else:
        #     q.put(('------------------------------')+('%s上%s更新失败\n' % (self.ip, self.rows['cpcontentid']))
        #           +(res.text)+('\n------------------------------\n'))
        #     # print('%s上%s更新失败' % (self.ip, self.rows['cpcontentid']))
        #     # print(res.text)
        #     # print('------------------------------')
        time.sleep(1)
        return res.text


def get_ch():
    tv_info = {}
    file_path = os.getcwd()+'\Modify_information.xlsx'
    print(file_path)
    wb = openpyxl.load_workbook(file_path)
    sh = wb.active
    rows = sh.max_row
    print(rows)
    for row in range(rows):
        tv_info[sh.cell(row=row+1, column=1).value] = sh.cell(row=row+1, column=2).value
    #print(tv_info)
    print(len(tv_info))
    return tv_info



def change(server, server_dict, tv_na_dict, tv_ori_dict, q):
    tv_list = server_dict[server]
    q.put(tv_list)
    # filepath = os.getcwd() + '\log.txt'
    a = Cms(server)
    q.put('登录cms：' + server)
    a.login()
    for ch in tv_list:
        print(ch)
        try:
            q.put('开始修改'+ch)
            a.select_ch(ch)
            status = a.updata_ch(tv_name=tv_na_dict[ch], url=None, origin=tv_ori_dict[ch])
            # q.put(status)
            if status == 'true':

                q.put('      在服务器'+server+'是否更新成功: '+status+'')

            else:

                q.put('      在服务器'+server+'更新失败 ')

            print('      在服务器'+server+'是否更新成功: '+status+'\n')
            # a.select_ch(ch)
            time.sleep(1)
        except (Exception, BaseException) as e:
        #     pass
            q.put(('%s 更新 %s 信息失败\n' % (server, ch))+('#########################################################\n')
                  +('str(Exception):\t\n', str(Exception))+('str(e):\t\t\n', str(e))+('repr(e):\t\n', repr(e))
                  +('##########################################################\n'))
            # print('e.message:\t', e.message)
            # print('traceback.print_exc():', traceback.print_exc())
            # print('traceback.format_exc():\n%s' % traceback.format_exc())


if __name__ == '__main__':

    q = multiprocessing.Manager().Queue()
    wb = openpyxl.Workbook()
    sh = wb.active
    filepath = os.getcwd() + '\log.txt'
    server_dict = {}
    server_list = []
    tv_na_dict = {}
    tv_ori_dict = {}
    # print(server_list)
    Tv_info = get_ch()
    ch_list = list(Tv_info.keys())

    b = MyDb()
    conn = b.connect_db()
    for TV_ch in ch_list:
        ip_list = b.get_info(conn, TV_ch)
        # print(TV_ch)
        # make server_dict
        for ip in ip_list:
            # print(type(ip))
            if ip in server_list:
                server_dict[ip].append(TV_ch)
            else:
                server_list.append(ip)
                server_dict.setdefault(ip, []).append(TV_ch)
                # print(server_list)
                # print(server_dict)
        tv_name = None
        origin = None
        c = Tv_info[TV_ch]
        tv_n = re.findall(r'(?<=tv_name=)[^,]+', c)

        if len(tv_n) != 0:
            tv_name = tv_n[0]
        tv_na_dict[TV_ch] = tv_name

        ori = re.findall(r'(?<=origin=)[^,]+', c)
        if len(ori) != 0:
            origin = ori[0]
        tv_ori_dict[TV_ch] = origin

    # print(server_list)
    # print(server_dict)
    # print(tv_na_dict)
    # print(tv_ori_dict)
    b.close(conn)

    # pool = multiprocessing.Pool(3)
    # # print(multiprocessing.cpu_count()-1)
    #
    # print('开始多进程修改')
    # results = []
    # for server in server_list:
    #     # print(server)
    #     # change(server, server_dict, tv_na_dict, tv_ori_dict)
    #     pool.apply_async(change, args=(server, server_dict, tv_na_dict, tv_ori_dict, q))
    # pool.close()
    # pool.join()

    for server in server_list:
        print(server)
        p = multiprocessing.Process(target=change, args=(server, server_dict, tv_na_dict, tv_ori_dict, q))
        p.daemon = True
        p.start()
        p.join()
    # change('212.8.243.177', server_dict, tv_na_dict, tv_ori_dict, q)
    print('q.qsize:')
    print(q.qsize(),"2222222222222")
    while not q.empty():
        print(q.get())
    print('多进程修改完毕')

