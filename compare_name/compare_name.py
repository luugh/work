# !/usr/bin/python3
# -*- coding: utf-8 -*-
'''
@File    :  compare_name.py  
@Software:  PyCharm
@Modify Time:   2019/10/23 13:32
@Author:    LGH
@Version    
@Desciption
------------      -------    --------    -----------
'''

import difflib
import openpyxl
import os


def get_name():
    path = os.getcwd()
    file_path = path + '/比较列表.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    cols = ws.max_column
    rows = ws.max_row
    name_exist = []
    name_new = []
    # for num in range(rows):
    #     name_exist.append(ws.cell(num+2, 1).value)
    name_exist = [ws.cell(num+2, 1).value for num in range(rows) if ws.cell(num+2, 1).value != None]
    # for num in range(rows):
    #     name_new.append(ws.cell(num+2, 2).value)
    name_new = [ws.cell(num+2, 2).value for num in range(rows) if ws.cell(num+2, 2).value != None]
    return name_exist, name_new


def compare_name(list1, list2):
    for n in list1:

        for m in list2:
            seq = difflib.SequenceMatcher(lambda x: x in " ", n.lower().rstrip(' hd'), m.lower().rstrip(' hd'))
            if seq.ratio() > 0.6:
                print(str(seq.ratio())+"---"+n+"---"+m)


if __name__ == "__main__":
    old_l, new_l = get_name()
    print(old_l)
    print(new_l)
    compare_name(old_l, new_l)