# !/usr/bin/python3
# -*- coding: utf-8 -*-

# @Time    : 2021/1/19 13:16
# @Author  : LGH
# @File    : ch_origin.py
# @Software: PyCharm

import openpyxl


class Origin:

    def get_data(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        rows = ws.max_row
        # print(rows)
        ori_dic = {}
        for n in range(1, rows+1):
            if ws.cell(n, 1).value is not None:
                ori_dic[ws.cell(n, 1).value] = ws.cell(n, 2).value

        return ori_dic

    def merge_data(self, dic1, dict2):
        wb = openpyxl.Workbook()
        ws = wb.active
        key_list = dic1.keys()
        n = 1
        for key in key_list:
            ws.cell(n, 1, key)
            ws.cell(n, 2, dic1.setdefault(key, '')+'_'+dict2.setdefault(key, ''))
            n += 1

        wb.save('origin_result.xlsx')


if __name__ == '__main__':
    a = Origin()
    ori_path = './origin.xlsx'
    lan_path = 'lan.xlsx'
    origin = a.get_data(ori_path)
    lan = a.get_data(lan_path)
    # print(origin)
    # print(lan)
    a.merge_data(origin, lan)