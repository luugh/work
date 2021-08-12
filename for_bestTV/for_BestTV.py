# !/usr/bin/python3
# -*- coding: utf-8 -*-

# @Time    : 2021/1/18 17:38
# @Author  : LGH
# @File    : for_BestTV.py
# @Software: PyCharm


import requests
import json
import openpyxl


class BestTV:

    def get_info(self,):
        url = 'http://cmc.best-tv.me:8080/panel_api.php?username=CHINA_ALV_LOCAL&password=l7HoCzqck4a2ceO&type=m3u&output=ts'
        res = requests.get(url)
        print(type(res.text))
        dict_li = json.loads(res.text)

        print(type(dict_li))
        # print('{}'.format(dict_l))

        dict_l = dict_li['available_channels']
        # print(dict_avil)
        print(len(dict_l.keys()))
        print(dict_l.keys())
        id_list = dict_l.keys()
        id_num = {}
        id_name = {}
        id_stream_type = {}
        id_type_name = {}
        id_stream_id = {}
        id_stream_icon = {}
        id_epg_channel_id ={}
        id_added ={}
        id_category_name = {}
        id_category_id={}
        id_series_no={}
        id_live={}
        id_container_extension={}
        id_custom_sid={}
        id_tv_archive={}
        id_direct_source={}
        id_tv_archive_duration={}
        for id in id_list:
            # print(id)
            # print(type(id))
            print(dict_l[id]['num'])
            id_num[id] = dict_l[id]['num']
            id_name[id] = dict_l[id]['name']
            id_stream_type[id] = dict_l[id]['stream_type']
            id_type_name[id] = dict_l[id]['type_name']
            id_stream_id[id] = dict_l[id]['stream_id']
            id_stream_icon[id] = dict_l[id]['stream_icon']
            id_epg_channel_id[id] = dict_l[id]['epg_channel_id']
            id_added[id] = dict_l[id]['added']
            id_category_name[id] = dict_l[id]['category_name']
            id_category_id[id] = dict_l[id]['category_id']
            id_series_no[id] = dict_l[id]['series_no']
            id_live[id] = dict_l[id]['live']
            id_container_extension[id] = dict_l[id]['container_extension']
            id_custom_sid[id] = dict_l[id]['custom_sid']
            id_tv_archive[id] = dict_l[id]['tv_archive']
            id_direct_source[id] = dict_l[id]['direct_source']
            id_tv_archive_duration[id] = dict_l[id]['tv_archive_duration']

        # print('{}'.format(id_name))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, 'ID')
        ws.cell(1, 2, 'name')
        ws.cell(1, 3, 'stream_id')
        # ws.cell(1, 5, 'stream_icon')
        ws.cell(1, 4, 'category_name')
        # ws.cell(1, 6, 'num')
        # ws.cell(1, 7, 'stream_type')
        # ws.cell(1, 8, 'type_name')
        # ws.cell(1, 9, 'epg_channel_id')
        # ws.cell(1, 10, 'added')
        # ws.cell(1, 11, 'category_id')
        # ws.cell(1, 12, 'series_no')
        # ws.cell(1, 13, 'live')
        # ws.cell(1, 14, 'container_extension')
        # ws.cell(1, 15, 'custom_sid')
        # ws.cell(1, 16, 'tv_archive')
        # ws.cell(1, 17, 'direct_source')
        # ws.cell(1, 18, 'tv_archive_duration')
        n = 0
        for id in id_list:
            ws.cell(n+2, 1, id)
            ws.cell(n+2, 2, id_name[id])
            ws.cell(n+2, 3, id_stream_id[id])
            # ws.cell(n+2, 5, id_stream_icon[id])
            ws.cell(n+2, 4, id_category_name[id])
            # ws.cell(n+2, 6, id_num[id])
            # ws.cell(n+2, 7, id_stream_type[id])
            # ws.cell(n+2, 8, id_type_name[id])
            # ws.cell(n+2, 9, id_epg_channel_id[id])
            # ws.cell(n + 2, 10, id_added[id])
            # ws.cell(n + 2, 11, id_category_id[id])
            # ws.cell(n+2, 12, id_series_no[id])
            # ws.cell(n+2, 13, id_live[id])
            # ws.cell(n+2, 14, id_container_extension[id])
            # ws.cell(n+2, 15, id_custom_sid[id])
            # ws.cell(n+2, 16, id_tv_archive[id])
            # ws.cell(n+2, 17, id_direct_source[id])
            # ws.cell(n+2, 18, id_tv_archive_duration[id])
            n += 1
        wb.save('./BestTV.xlsx')
        print('结果已写入表格')

if __name__ == "__main__":
    a = BestTV()
    a.get_info()