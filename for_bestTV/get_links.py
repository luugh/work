# !/usr/bin/python3
# -*- coding: utf-8 -*-

# @Time    : 2021/1/27 16:56
# @Author  : LGH
# @File    : get_links.py
# @Software: PyCharm


import openpyxl


def get_all():
    w = openpyxl.load_workbook('./bestTV_data.xlsx')
    sh = w.active
    ch_list = []
    ch_name = {}
    ch_link = {}
    ch_orig = {}
    rows = sh.max_row
    for row in range(2, rows+1):
        ch_list.append(sh.cell(row, 1).value)
        ch_name[sh.cell(row, 1).value] = sh.cell(row, 2).value
        ch_link[sh.cell(row, 1).value] = sh.cell(row, 3).value
        ch_orig[sh.cell(row, 1).value] = sh.cell(row, 4).value

    w2 = openpyxl.load_workbook('./target.xlsx')
    sh2 = w2.active
    w3 = openpyxl.Workbook()
    sh3 = w3.active
    target_list = []
    rows2 = sh2.max_row

    for row in range(1, rows2+1):
        sh3.cell(row, 1, sh2.cell(row, 1).value)
        sh3.cell(row, 2, ch_name.setdefault(sh2.cell(row, 1).value, ''))
        sh3.cell(row, 3, ch_link.setdefault(sh2.cell(row, 1).value, ''))
        sh3.cell(row, 4, ch_orig.setdefault(sh2.cell(row, 1).value, ''))

    w3.save('./target_links.xlsx')



    print(len(ch_list))


if __name__ == '__main__':
    get_all()