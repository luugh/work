# !/usr/bin/python3
# -*- coding: utf-8 -*-

# @Time    : 2021/1/30 9:54
# @Author  : LGH
# @File    : panel.q.py
# @Software: PyCharm


from lxml import etree
from selenium import webdriver
import openpyxl
import time


def get_s():
    url = 'https://panel.q-hd.net/fastprinter.php'
    headers = {
        'Accept': 'text/html, application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'Accept-Encoding': 'gzip,deflate,br',
        'Accept-Language': 'zh-CN',
        'Connection': 'keep-alive',
        'Host': 'panel.q-hd.net',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0(Windows NT 10.0;Win64;x64;rv: 86.0) Gecko/20100101 Firefox/86.0'
    }
    # res = requests.get(url=url, headers=headers)
    # dom = etree.HTML(res.text)



    # driver = webdriver.Chrome()
    driver = webdriver.Firefox()
    driver.set_window_size(1280, 720)
    driver.set_page_load_timeout(10)
    driver.set_script_timeout(30)

    try:
        a = time.time()
        # driver.get(url)
    except Exception as e:
        print('except加载时间已到10s')
        driver.execute_script('window.stop()')
    finally:
        print('finally加载时间已到10s')
        html = driver.page_source

    print('window.stop()')
    b=time.time()

    html = driver.page_source
    print(html)
    print('花费时间：{}'.format(b - a))

    # dom = etree.parse(r'E:\untitled3\for_panel.q-hd\panel.html', etree.HTMLParser())

    with open(r'html.txt', 'a', encoding='utf-8') as f:
        f.write(html)
    # print()
    driver.quit()


def use_html():
    dom = etree.parse(r'E:\untitled3\for_panel.q-hd\panel.html', etree.HTMLParser())
    cate_ch = {}
    ch_list = []
    category = []
    # print(dom.xpath('/html/body/center/div[2]/h3/text()'))

    # print(dom.xpath('/html/body/center/div[2]/div[3]/div/div/text()'))
    # /html/body/center/div[2]/div[2]/div/div
    # /html/body/center/div[2]/div[4]/div/div

    for i in range(1, 100):
        print('/html/body/center/div[{}]/h3/text()'.format(i))
        # print()
        resu = dom.xpath('/html/body/center/div[{}]/h3/text()'.format(i))
        if len(resu) != 0:
            print(resu)
        for n in range(1, 500):
            # print('/html/body/center/div[{}]/div[{}]/div/div/text()'.format(i, n))
            resu1 = dom.xpath('/html/body/center/div[{}]/div[{}]/div/div/text()'.format(i, n))
            # print(dom.xpath('/html/body/center/div[{}]/div[{}]/div/div/text()'.format(i, n)))
            if len(resu1) != 0:
                print(resu1)
                ch_list.append(resu1[0])
        if len(resu) != 0:
            category.append(resu[0])
            # print(category)
            # print(ch_list)
            cate_ch[resu[0]] = tuple(ch_list)
            ch_list = []

    print(cate_ch)
    print(len(cate_ch))
    print(cate_ch.keys())
    print(category)

    with open(r'E:\untitled3\for_panel.q-hd\result.txt', 'w+', encoding='utf-8') as f:
        for cate in category:
            f.write('{}\n'.format(cate))
            print(cate)
            for ch in cate_ch[cate]:
                print(ch)
                f.write('{}\n'.format(ch))
            f.write('\n')
            f.write('**********************************')
    w1 = openpyxl.Workbook()
    for name in cate_ch.keys():
        sh = w1.create_sheet(name)
        nums = len(cate_ch[name])
        for n in range(nums):
            print(type(cate_ch[name][n]))
            print(cate_ch[name][n].encode('utf-8'))
            sh.cell(n+1, 1, cate_ch[name][n].encode('utf-8'))
    w1.save(r'./result.xlsx')


if __name__ == '__main__':
    # get_s()
    use_html()