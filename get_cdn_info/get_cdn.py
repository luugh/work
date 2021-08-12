#!/usr/bin/python3
# -*- coding:utf-8 -*-
# by lgh
# change channel's name

from change_name.batch_modification import MyDb
import pymysql
import os
import openpyxl


def get_ch():
    tv_info = []
    file_path = os.getcwd() + '\Modify_information.xlsx'
    print(file_path)
    wb = openpyxl.load_workbook(file_path)
    sh = wb.active
    rows = sh.max_row
    print(rows)
    for row in range(rows):
        tv_info.append(sh.cell(row=row+1, column=1).value)
    print(tv_info)
    print(len(tv_info))

    return tv_info


def create_excl(ch_list, ch_dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    # print(ch_list)
    print('结果写入表格：result.xlsx')
    num = 1
    for ch in ch_list:
        ws.cell(num, 1, ch)
        ws.cell(num, 2, ch_dict[ch])
        num = num + 1
    file_path = os.getcwd() + r'\result.xlsx'
    wb.save(file_path)


if __name__ == '__main__':
    wb = openpyxl.Workbook()
    sh = wb.active
    TV_info = get_ch()
    ch_list = TV_info
    print(ch_list)

    b = MyDb()
    conn = b.connect_db()
    ch_dict = {}
    for TV_name in ch_list:
        Ip_list = b.get_info(conn, TV_name)

        if len(Ip_list) != 0:
            ch_dict[TV_name] = Ip_list[0]
        else:
            ch_dict[TV_name] = '未找到cdn信息'

    b.close(conn)
    create_excl(ch_list, ch_dict)
