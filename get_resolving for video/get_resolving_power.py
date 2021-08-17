#! /usr/bin/python3
# -*-coding:utf-8-*-
# by lgh
# @Time     :2019/3/19 9:41
# @Author   :LGH
# @Site     :
# @File     :get_resolving_power.py
# @Software :
import os
import sys
import shutil
import json
import subprocess
import openpyxl


def get_ch():
    tv_info = {}
    tv_name =[]
    file_path = os.getcwd()+'\ch_info.xlsx'
    # print(file_path)
    wb = openpyxl.load_workbook(file_path)
    sh = wb.active
    rows = sh.max_row
    # print(rows)
    for row in range(rows):
        tv_name.append(sh.cell(row=row+1, column=1).value)
        tv_info[sh.cell(row=row+1, column=1).value] = sh.cell(row=row+1, column=2).value
    # print(tv_name)
    # print(tv_info)
    # print(len(tv_info))
    return tv_name, tv_info


def create_excl(name_list, tv_info, tv_resolving, Encoder, lan):
    wb = openpyxl.Workbook()
    ws = wb.active
    print('结果写入表格：result.xlsx')
    for num in range(0, len(name_list)):
        ws.cell(num+1, 1, name_list[num])
        print(name_list[num])
        ws.cell(num+1, 2, tv_info[name_list[num]])
        print(tv_info[name_list[num]])
        print(tv_resolving[name_list[num]])
        ws.cell(num+1, 3, tv_resolving[name_list[num]])
        ws.cell(num+1, 4, Encoder[name_list[num]])
        ws.cell(num+1, 5, lan[name_list[num]])
    file_path = os.getcwd() + r'\result.xlsx'
    wb.save(file_path)


def get_resolving(url):
    command = 'ffprobe -loglevel quiet -print_format json -show_format -show_streams -i ' + url
    # print(command)
    result = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    out = result.stdout.read()
    # print(str(out))
    temp = str(out.decode('utf-8'))
    # print(temp)
    try:
        data = str(json.loads(temp)['streams'][0]['width'])+'x'+str(json.loads(temp)['streams'][0]['height'])
        print(data)
        date_Encoder = str(json.loads(temp)['streams'][0].get('codec_name', 'null'))
        date_language = str(json.loads(temp).get('tags', 'null'))
    except:
        data = str(json.loads(temp))
        print(data)
    return data, date_Encoder, date_language


if __name__ == '__main__':
    tv_resolving = {}
    Encoder = {}
    lan = {}
    tv_name, tv_info = get_ch()
    # url = 'http://212.8.243.139:8082/TV52127'
    for name in tv_name:
        print('正在获取'+name+'的分辨率')
        url = tv_info[name]
        resolving, Encoder_date, lan_date = get_resolving(tv_info[name])
        tv_resolving[name] = resolving
        Encoder[name] = Encoder_date
        lan[name] = lan_date

    create_excl(tv_name, tv_info, tv_resolving, Encoder, lan)
