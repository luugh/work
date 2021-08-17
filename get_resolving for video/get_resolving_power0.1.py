#!/usr/bin python3
# -*- encoding: utf-8 -*-
'''
@File    :   get_resolving_power0.1.py   
@Contact :   
@License :   

@Modify Time      @Author    @Version    @Desciption
------------      -------    --------    -----------
2019/5/5 22:21    lgh        1.0         更新使用ffmpeg-python模块，不再单独使用ffmpeg程序
'''

import os
# install ffmpeg-python
import ffmpeg
import json
import subprocess
import openpyxl


def get_ch():
    tv_info = {}
    tv_name =[]
    #windows
    # file_path = os.getcwd()+'\ch_info.xlsx'
    #unix
    file_path = os.getcwd() + '/ch_info.xlsx'
    # print(file_path)
    wb = openpyxl.load_workbook(file_path)
    sh = wb.active
    rows = sh.max_row
    # print(rows)
    for row in range(rows):
        tv_name.append(sh.cell(row=row+1, column=1).value)
        tv_info[sh.cell(row=row+1, column=1).value] = sh.cell(row=row+1, column=2).value
    # print(tv_name)
    # print(tv_info)
    # print(len(tv_info))
    return tv_name, tv_info


def create_excl(name_list, tv_info, tv_resolving, Encoder, lan):
    wb = openpyxl.Workbook()
    ws = wb.active
    print('结果写入表格：result.xlsx')
    for num in range(0, len(name_list)):
        ws.cell(num+1, 1, name_list[num])
        print(name_list[num])
        ws.cell(num+1, 2, tv_info[name_list[num]])
        print(tv_info[name_list[num]])
        print(tv_resolving[name_list[num]])
        ws.cell(num+1, 3, tv_resolving[name_list[num]])
        ws.cell(num+1, 4, Encoder[name_list[num]])
        ws.cell(num+1, 5, lan[name_list[num]])
    #windows
    # file_path = os.getcwd() + r'\result.xlsx'
    #unix
    file_path = os.getcwd() + r'/result.xlsx'
    wb.save(file_path)


def get_resolving(url):
    # command = 'ffprobe -loglevel quiet -print_format json -show_format -show_streams -i ' + url
    # # print(command)
    # result = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    # out = result.stdout.read()
    # # print(str(out))
    # temp = str(out.decode('utf-8'))
    # # print(temp)
    try:

        temp = ffmpeg.probe(url)
        # data = str(json.loads(temp)['streams'][0]['width'])+'x'+str(json.loads(temp)['streams'][0]['height'])
        # print(data)
        # date_Encoder = str(json.loads(temp)['streams'][0].get('codec_name', 'null'))
        # date_language = str(json.loads(temp).get('tags', 'null'))
        stream_list = temp['streams']
        # print(stream_list)
        stream_len = len(temp['streams'])
        for i in range(stream_len):
            codec_type = stream_list[i].get('codec_type', 'null')
            if codec_type == 'video':
                data = str(stream_list[i]['width']) + 'x' + str(stream_list[i]['height'])
                data_Encoder = str(stream_list[i].get('codec_name', 'null'))
                data_language = str(stream_list[i].get('tags', 'null'))
    except:
        data = 'N/A'
        data_Encoder = 'N/A'
        data_language = 'N/A'
        # print(data)
    return data, data_Encoder, data_language


if __name__ == '__main__':
    tv_resolving = {}
    Encoder = {}
    lan = {}
    tv_name, tv_info = get_ch()
    # url = 'http://212.8.243.139:8082/TV52127'
    for name in tv_name:
        print('正在获取'+name+'的分辨率').decode('')
        url = tv_info[name]
        resolving, Encoder_date, lan_date = get_resolving(tv_info[name])
        tv_resolving[name] = resolving
        Encoder[name] = Encoder_date
        lan[name] = lan_date

    create_excl(tv_name, tv_info, tv_resolving, Encoder, lan)
