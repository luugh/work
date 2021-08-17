#! /usr/bin/python3
# -*-coding:utf-8-*-
# by lgh
# @Time     :2019/3/19 9:41
# @Author   :LGH
# @Site     :
# @File     :get_resolving_power.py
# @Software :
import os
import sys
import shutil
import json
import subprocess
import openpyxl
import eventlet
import re


def get_dir():
    with open(r'./root_dir.txt') as f:
        lines = f.readlines()
    root_dir = []
    for line in lines:
        root_dir.append(line.strip(r'\n'))
    print '正在获取视频路径'
    file_path_name_dict = {}
    for root_path in root_dir:
        for root, dirs, files in os.walk(root_path):
            for file in files:
                file_path_name_dict[file] = os.path.join(root, file)
    print '获取完成'
    return file_path_name_dict


def create_excl(name_list,file_paths ,tv_resolving, Encoder, bitrate, subtitle):
    wb = openpyxl.Workbook()
    ws = wb.active
    print('结果写入表格：result.xlsx')
    ws.cell(1, 1, '视频名称')
    ws.cell(1, 2, '路径')
    ws.cell(1, 3, '分辨率')
    ws.cell(1, 4, '视频编码')
    ws.cell(1, 5, '码率')
    ws.cell(1, 6, '是否有字幕')
    for num in range(1, len(name_list)):
        ws.cell(num+1, 1, name_list[num])
        # print(name_list[num])
        # print(tv_resolving[name_list[num]])
        ws.cell(num+1, 2, file_paths[name_list[num]])
        ws.cell(num+1, 3, str(tv_resolving[name_list[num]]))
        ws.cell(num+1, 4, Encoder[name_list[num]])
        # ws.cell(num+1, 5, lan[name_list[num]])
        # print bitrate[name_list[num]]
        ws.cell(num+1, 5, bitrate[name_list[num]])
        ws.cell(num+1, 6, subtitle[name_list[num]])
    file_path = os.getcwd() + r'/result.xlsx'   #unix

    wb.save(file_path)
    print '写入表格完成'


def get_resolving(url):
    print url
    # command = '/opt/OceanView/CDN/lts/lib/trans_probe -loglevel quiet -print_format json -show_format -show_streams -i ' + url
    # command = '/opt/starview/cdn/lts/lib/lts_probe -loglevel quiet -print_format json -show_format -show_streams -i ' + url
    command = './lts_probe -loglevel quiet -print_format json -show_format -show_streams -i "' + url + '"'
    command1 = './lts_probe -show_format -i "' + url + '"'
    # print(command)
    key_num = 0
    eventlet.monkey_patch()
    with eventlet.Timeout(30, False):
        result = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        out = result.stdout.read()
        result = subprocess.Popen(command1, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        out1 = result.stdout.read()
    # print(str(out))
        temp = str(out)
        temp1 = str(out1)
        # print temp
        # print temp1
        key_num = 1
    # print(temp)
    if key_num == 0:
        temp = str({'none': ''})
    # print(temp)
    # print(type(temp))
    try:
        # 获取码率
        bit_rate_list = re.findall(r'(?<=bitrate:).+', temp1)
        # print bit_rate_list
        bit_rate = bit_rate_list[0].strip()
        print bit_rate
        # 获取分辨率，编码
        data_list = json.loads(temp)['streams']
        # print data_list
        for n in data_list:
            key_list = n.keys()
            # print(key_list)
            if 'width' in key_list and n['codec_type'] == 'video':
                # print(n['width'])
                # print(n['height'])
                data = str(n['width'])+'x'+str(n['height'])
                print(data)
                # data = str(n['height'])+'p'
                data_encoder = str(n.get('codec_name', 'N/A'))
                data_language = str(n.get('tags', 'N/A'))
            if 'coded_width' in key_list and n['codec_type'] == 'video':
                date = str(n['coded_width'])+'x'+str(n['coded_height'])
                data_encoder = str(n.get('codec_name', 'N/A'))
            if n['codec_type'] == 'subtitle':
                data_subtitle = 'yes'
            else:
                data_subtitle = 'no'
        # data = str(json.loads(temp)['streams'][0]['width'])+'x'+str(json.loads(temp)['streams'][0]['height'])
        # print(data)
        # data_Encoder = str(json.loads(temp)['streams'][0].get('codec_name', 'N/A'))
        # print(data_Encoder)
        # data_language = str(json.loads(temp).get('tags', 'N/A'))
        # print(data_language)
    except:
        data = temp
        data_encoder = 'N/A'
        data_language = 'N/A'
        data_subtitle = 'N/A'
        bit_rate = 'N/A'
        # print(data)
    # return data, data_encoder, data_language, data_subtitle
    return data, data_encoder, bit_rate,data_subtitle


if __name__ == '__main__':
    tv_resolving = {}
    Encoder = {}
    lan = {}
    subtitle = {}
    # dir_lines = get_ch()
    dir_name_dict = get_dir()
    names = []
    file_path = {}
    bitrate = {}
    # print dir_lines
    print '开始获取分辨率，码率信息'
    for name in dir_name_dict.keys():
        names.append(name)
        url = dir_name_dict[name]
        resolving, Encoder_data, bit_rate,subtitle_data = get_resolving(url)
        tv_resolving[name] = resolving
        Encoder[name] = Encoder_data
        # lan[name] = lan_data
        bitrate[name] = bit_rate
        subtitle[name] = subtitle_data
        file_path[name] = url
    # print file_path
    print '提取信息完成'
    # create_excl(names, file_path, tv_resolving, Encoder, lan, subtitle)
    create_excl(names, file_path, tv_resolving, Encoder, bitrate,subtitle)