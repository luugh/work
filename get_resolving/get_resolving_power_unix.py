#! /usr/bin/python3
# -*-coding:utf-8-*-
# by lgh
# @Time     :2019/3/19 9:41
# @Author   :LGH
# @Site     :
# @File     :get_resolving_power.py
# @Software :
import os
import sys
import shutil
import json
import subprocess
import openpyxl
import eventlet



def get_ch():
    tv_info = {}
    tv_name =[]
    file_path = os.getcwd()+'/ch_info.xlsx'
    print(file_path)
    wb = openpyxl.load_workbook(file_path)
    sh = wb.active
    rows = sh.max_row
    # print(rows)
    for row in range(rows):
        tv_name.append(sh.cell(row=row+1, column=1).value)
        tv_info[sh.cell(row=row+1, column=1).value] = sh.cell(row=row+1, column=2).value
    # print(tv_name)
    # print(tv_info)
    # print(len(tv_info))
    return tv_name, tv_info


def create_excl(name_list, tv_info, tv_resolving, Encoder, lan, subtitle):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, '唯一标识')
    ws.cell(1, 2, 'links')
    ws.cell(1, 3, '分辨率')
    ws.cell(1, 4, '编码')
    ws.cell(1, 5, '语言')
    ws.cell(1, 6, '字幕')
    print('write the results to excle: result.xlsx')
    for num in range(0, len(name_list)):
        ws.cell(num+2, 1, name_list[num])
        print(name_list[num])
        ws.cell(num+2, 2, tv_info[name_list[num]])
        print(tv_info[name_list[num]])
        print(tv_resolving[name_list[num]])
        ws.cell(num+2, 3, str(tv_resolving[name_list[num]]))
        ws.cell(num+2, 4, Encoder[name_list[num]])
        ws.cell(num+2, 5, lan[name_list[num]])
        ws.cell(num+2, 6, subtitle[name_list[num]])
    file_path = os.getcwd() + r'/result.xlsx'
    wb.save(file_path)


def get_resolving(url):
    file_path = os.getcwd()
    # command = '/opt/OceanView/CDN/lts/lib/trans_probe -loglevel quiet -print_format json -show_format -show_streams -i ' + url
    # command = '/opt/starview/cdn/lts/lib/lts_probe -loglevel quiet -print_format json -show_format -show_streams -i ' + url
    command = file_path + '/lts_probe -loglevel quiet -print_format json -show_format -show_streams -i ' + url
    # print(command)
    key_num = 0
    eventlet.monkey_patch()
    with eventlet.Timeout(30, False):
        result = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        out = result.stdout.read()
    # print(str(out))
        temp = str(out.decode('utf-8'))
        key_num = 1
    # print(temp)
    if key_num == 0:
        temp = str({'none': ''})
    print(temp)
    print(type(temp))
    try:
        data_list = json.loads(temp)['streams']
        print('temp:{}'.format(temp))
        data = 'N/A'
        data_encoder = 'N/A'
        data_language = 'N/A'
        data_subtitle = 'N/A'
        for n in data_list:
            key_list = n.keys()
            # print(key_list)
            if 'width' in key_list and n['codec_type'] == 'video':
                print(n['width'])
                print(n['height'])
                data = str(n['width'])+'x'+str(n['height'])
                print(data)
                # data = str(n['height'])+'p'
                data_encoder = str(n.get('codec_name', 'N/A'))
                data_language = str(n.get('tags', 'N/A'))
            if n['codec_type'] == 'subtitle':
                data_subtitle = 'yes'
            else:
                data_subtitle = 'no'
        # data = str(json.loads(temp)['streams'][0]['width'])+'x'+str(json.loads(temp)['streams'][0]['height'])
        # print(data)
        # data_Encoder = str(json.loads(temp)['streams'][0].get('codec_name', 'N/A'))
        # print(data_Encoder)
        # data_language = str(json.loads(temp).get('tags', 'N/A'))
        # print(data_language)
    except:
        if len(temp) == 0:
            data = 'N/A'
        else:
            data = temp
        data_encoder = 'N/A'
        data_language = 'N/A'
        data_subtitle = 'N/A'
    print('data: {}'.format(data))

    return data, data_encoder, data_language, data_subtitle


if __name__ == '__main__':
    tv_resolving = {}
    Encoder = {}
    lan = {}
    subtitle = {}
    tv_name, tv_info = get_ch()
    # url = 'http://xxxxxx:8082/TV52127'
    nums = len(tv_name)
    n = 1
    for name in tv_name:
        print('geting '+str(name)+"'resolving")
        url = tv_info[name]
        resolving, Encoder_data, lan_data, subtitle_data = get_resolving(tv_info[name])
        tv_resolving[name] = resolving
        Encoder[name] = Encoder_data
        lan[name] = lan_data
        subtitle[name] = subtitle_data
        print('已完成：{}/{}'.format(n, nums))
        n += 1

    create_excl(tv_name, tv_info, tv_resolving, Encoder, lan, subtitle)
