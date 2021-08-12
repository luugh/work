# !/usr/bin/python3
# -*- coding: utf-8 -*-

# @Time    : 2021/1/25 17:55
# @Author  : LGH
# @File    : Kingofsat.py
# @Software: PyCharm


import os
import time
import requests
import openpyxl
from lxml import etree
from bs4 import BeautifulSoup


class GetSat:

    def __init__(self):
        self.session = ''
        self.dom = ''

    def get_dom(self, ch_name):
        # ch_name = 'A Spor HD'
        url = 'https://en.kingofsat.net/find.php?question=' + ch_name
        res = requests.get(url)
        html = res.content
        res.close()
        # print('html:{}'.format(html))
        # self.dom = etree.HTML(html)
        self.dom = BeautifulSoup(html, "html.parser")
        self.dom.prettify()

    def get_info(self, cp_name, sat_info):
        # 卫星：频点：SID
        # sat_info:bld:s
        # /html/body/div[5]/table[9]/tbody/tr/td[@class='pos'] 42.0°E
        # /html/body/div[5]/table[9]/tbody/tr/td[@class='bld'] 12054.00, H
        #
        # table1 = self.dom.xpath('/html/body/div[@class="w3-main"]/table[@class="frq"]')

        # table1 = self.dom.xpath('/html/body/div[@class="w3-main"]/table[1]/tr/td[@class="pos"]')
        # print('table1:{}'.format(table1))

        tables = self.dom.find_all('table')
        # .next_siblings生成迭代器对象
        # table1 = self.dom.find('table').next_sibling.next_sibling
        fre_p = None
        ch_name = None
        sid = None
        # cp_name = cp_name.replace('ü', 'u')
        for table1 in tables:

            if table1.find('td', class_='pos') and table1.find('td', class_='pos').string == sat_info:
                # print('table1:{}'.format(table1))
                print(table1.find('td', class_='pos').string)
                # 节目信息在div下的tr里面
                div = table1.next_sibling.next_sibling
                trs = div.find_all('tr')
                # print('trs:{}'.format(trs))
                for tr in trs:
                    print(tr.find('a', class_='A3'))
                    # print(cp_name.lower())
                    # print(tr.find('a', class_='A3').string.lower())
                    # print(cp_name.lower() in tr.find('a', class_='A3').string.lower())
                    if tr.find('a', class_='A3'):
                        if 'ü' in cp_name and 'ü' not in tr.find('a', class_='A3').string:
                            cp_name = cp_name.replace('ü', 'u')
                        elif 'ü' not in cp_name and 'ü' in tr.find('a', class_='A3').string:
                            cp_name = cp_name.replace('u', 'ü')
                        else:
                            pass
                        if tr.find('a', class_='A3') and cp_name.lower() in tr.find('a', class_='A3').string.lower():
                            if table1.find('td', class_='bld'):
                                fre_p = table1.find('td', class_='bld').string
                            elif table1.find('td', class_='nbld'):
                                fre_p = table1.find('td', class_='nbld').string
                            ch_name = tr.find('a', class_='A3').string
                            sid = tr.find('td', class_='s').string
                            break
        if ch_name:
            print('ch_name:{}'.format(ch_name))
            print('fre_p:{}'.format(fre_p))
            print('sid:{}'.format(sid))


                # print('fre_p:{}'.format(fre_p))
                # print('sid:{}'.format(sid))
                # print('ch_name:{}'.format(ch_name))

        return ch_name, fre_p, sid


def get_ch():
    cp_ch = {}
    cp_sat = {}

    w1 = openpyxl.load_workbook('ch.xlsx')
    s1 = w1.active
    rows = s1.max_row
    for n in range(1, rows+1):
        if s1.cell(n, 1).value is not None:
            cp= s1.cell(n, 1).value
            cp_ch[cp] = s1.cell(n, 2).value
            cp_sat[cp] = s1.cell(n, 3).value

    return cp_ch, cp_sat


if __name__ == '__main__':

    cp_ch, cp_sat = get_ch()
    cp_list = cp_ch.keys()
    a = GetSat()
    ch_name_d = {}
    fre_p_d = {}
    sid_d = {}
    for cp in cp_list:
        print('******************')
        print('cp:{},cp_name:{}'.format(cp, cp_ch[cp]))
        try:
            a.get_dom(cp_ch[cp])
        except requests.exceptions.RequestException as e:
            print(e)

        ch_name = None
        fre_p = None
        sid = None
        ch_name, fre_p, sid = a.get_info(cp_ch[cp], cp_sat[cp])
        ch_name_d[cp] = ch_name
        fre_p_d[cp] = fre_p
        sid_d[cp] = sid
        time.sleep(2)
    for cp in cp_list:
        print('ch_name:{},fre:{},sid:{}'.format(ch_name_d[cp], fre_p_d[cp], sid_d[cp]))
    w2 = openpyxl.Workbook()
    s2 = w2.active
    nums = len(cp_list)
    # 标题
    s2.cell(1, 1, '唯一标识')
    s2.cell(1, 2, '节目名称')
    s2.cell(1, 3, '查询到的名称')
    s2.cell(1, 4, '卫星')
    s2.cell(1, 5, '频点信息')
    for i in range(1, nums+1):
        cp = list(cp_list)[i-1]
        s2.cell(i+1, 1, cp)
        s2.cell(i+1, 2, cp_ch[cp])
        s2.cell(i+1, 3, str(ch_name_d[cp]))
        s2.cell(i+1, 4, cp_sat[cp])
        s2.cell(i+1, 5, '{}:{}:{}'.format(cp_sat[cp], fre_p_d[cp], sid_d[cp]))
    w2.save('res.xlsx')