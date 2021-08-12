#!/usr/bin/python3
# -*- coding: utf-8 -*-

# @Time    : 2021/2/24 14:36
# @Author  : LGH
# @File    : tongji.py
# @Software: PyCharm

import openpyxl


def tong_ji(filename):
    w1 = openpyxl.load_workbook(filename)
    s1= w1.active
    rows = s1.max_row
    dict_ori = {}
    dict_name = {}
    dict_c1={}
    dict_c1a = {}
    cp_l = []
    for row in range(1,rows+1):
        cp = s1.cell(row, 3).value
        cp_l.append(cp)
        dict_ori[cp]=s1.cell(row,1).value
        dict_name[cp]=s1.cell(row,2).value
        dict_c1[cp] = s1.cell(row,4).value
        dict_c1a[cp] = s1.cell(row,5).value

    return cp_l,dict_ori,dict_name,dict_c1,dict_c1a


if __name__ == '__main__':
    cp_l1,dict_ori1,dict_name1,dict_c1,dict_c1a = tong_ji('1.04.xlsx')
    cp_l2,dict_ori2,dict_name2,dict_c2,dict_c2a = tong_ji('1.25.xlsx')
    w3 = openpyxl.Workbook()
    s3 = w3.active
    nums = len(cp_l1)
    n =1
    for cp in cp_l1:
        if cp in cp_l2:
            s3.cell(n,1,dict_ori1[cp])
            s3.cell(n,2,dict_name1[cp])
            s3.cell(n,3,cp)
            s3.cell(n,4,dict_c1[cp])
            s3.cell(n, 5, dict_c1a[cp])
            s3.cell(n, 6, dict_c2[cp])
            s3.cell(n, 7, dict_c2a[cp])
        else:
            s3.cell(n, 1, dict_ori1[cp])
            s3.cell(n, 2, dict_name1[cp])
            s3.cell(n, 3, cp)
            s3.cell(n, 4, dict_c1[cp])
            s3.cell(n, 5, dict_c1a[cp])
        n+=1
    w3.save('res.xlsx')