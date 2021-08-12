# !/usr/bin/python3
# -*- coding: utf-8 -*-
'''
@File    :  wangyi-epg.py  
@Software:  PyCharm
@Modify Time:   2019/8/16 9:38
@Author:    LGH
@Version    
@Desciption
------------      -------    --------    -----------
'''

import datetime
import requests
import time
import os
import openpyxl
from lxml import etree
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl.styles import  PatternFill


def init_web_driver():
    global driver
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    dirver_path = r'E:\untitled3\chromedriver\chromedriver.exe'
    driver = webdriver.Chrome(chrome_options=chrome_options, executable_path=dirver_path)


def close_web_driver():
    driver.quit()


def get_data(date_st, wb):
    # url = 'http://www.espn.com.br/programacao?date=' + '2019-08-17'
    url = 'http://www.espn.com.br/programacao?date=' + date_st
    driver.get(url)
    driver.implicitly_wait(20)
    # time.sleep(30)
    # print(driver.title)
    html = driver.page_source
    # print(html)
    Selector = etree.HTML(html)
    espn_b = []
    espn_b_d = {}
    espn = []
    espn_d = {}
    espn_e = []
    espn_e_d = {}
    espn_2 = []
    espn_2_d ={}
    # 提取espn-brasil
    for num in range(1, 100):
        # result 为一个列表
        Xpath_time = '//*[@data-channel="espn-brasil"]/div/article[%d]/time/text()' % num
        result_time = Selector.xpath(Xpath_time)
        # print(Selector.xpath(Xpath_time))
        if len(result_time) == 0:
            break
        start_time = result_time[0]
        Xpath_title = '//*[@data-channel="espn-brasil"]/div/article[%d]/span/text()' % num
        result_title = Selector.xpath(Xpath_title)
        title = result_title[0]
        espn_b.append(start_time)
        espn_b_d[start_time] = title
    # print(espn_b)
    # print(espn_b_d)
    for num in range(1, 100):
        # result 为一个列表
        Xpath_time = '//*[@data-channel="espn"]/div/article[%d]/time/text()' % num
        result_time = Selector.xpath(Xpath_time)
        # print(Selector.xpath(Xpath_time))
        if len(result_time) == 0:
            break
        start_time = result_time[0]
        Xpath_title = '//*[@data-channel="espn"]/div/article[%d]/span/text()' % num
        result_title = Selector.xpath(Xpath_title)
        title = result_title[0]

        espn.append(start_time)
        espn_d[start_time] = title
    # print(espn)
    # print(espn_d)
    for num in range(1, 100):
        # result 为一个列表
        Xpath_time = '//*[@data-channel="espn-extra"]/div/article[%d]/time/text()' % num
        result_time = Selector.xpath(Xpath_time)
        # print(Selector.xpath(Xpath_time))
        if len(result_time) == 0:
            break
        start_time = result_time[0]
        Xpath_title = '//*[@data-channel="espn-extra"]/div/article[%d]/span/text()' % num
        result_title = Selector.xpath(Xpath_title)
        title = result_title[0]

        espn_e.append(start_time)
        espn_e_d[start_time] = title
    # print(espn_e)
    # print(espn_e_d)

    for num in range(1, 100):
        # result 为一个列表
        Xpath_time = '//*[@data-channel="espn-2"]/div/article[%d]/time/text()' % num
        result_time = Selector.xpath(Xpath_time)
        # print(Selector.xpath(Xpath_time))
        if len(result_time) == 0:
            break
        start_time = result_time[0]
        Xpath_title = '//*[@data-channel="espn-2"]/div/article[%d]/span/text()' % num
        result_title = Selector.xpath(Xpath_title)
        title = result_title[0]

        espn_2.append(start_time)
        espn_2_d[start_time] = title
    # print(espn_2)
    # print(espn_2_d)

    ws = wb.create_sheet(title=date_st)
    write_excle(espn_2, espn_2_d, ws=ws, row=1, channel_name='espn 2', date_str=date_st)
    write_excle(espn, espn_d, ws=ws, row=2, channel_name='espn', date_str=date_st)
    write_excle(espn_b, espn_b_d, ws=ws, row=3, channel_name='espn brasil', date_str=date_st)
    write_excle(espn_e, espn_e_d, ws=ws, row=4, channel_name='espn extra', date_str=date_st)


def write_excle(list1, dict1, ws, row, channel_name, date_str):

    cols_num = len(list1)
    ws.cell(row, 1, channel_name)
    for num in range(cols_num):
        if num+1 < cols_num:
            ws.cell(row, num+2, list1[num]+'-'+list1[num+1]+' '+dict1[list1[num]]+' [' + channel_name+' '+date_str+' ' +
                    list1[num]+'-'+list1[num+1]+']')
            if 'PREMIER LEAGUE' in dict1[list1[num]] or 'CAMPEONATO ESPANHOL' in dict1[list1[num]]:
                ws.cell(row, num+2).fill = PatternFill('solid', fgColor='EEEE00')
            if '- AO VIVO' in dict1[list1[num]] and 'PREMIER LEAGUE' in dict1[list1[num]]:
                ws.cell(row, num + 2).fill = PatternFill('solid', fgColor='9F79EE')
            if '- AO VIVO' in dict1[list1[num]] and 'CAMPEONATO ESPANHOL' in dict1[list1[num]]:
                ws.cell(row, num + 2).fill = PatternFill('solid', fgColor='9F79EE')
        else:
            ws.cell(row, num+2, list1[num]+' '+dict1[list1[num]]+' [' + channel_name+' '+date_str+' '+
                    list1[num]+']')
            if 'PREMIER LEAGUE' in dict1[list1[num]] or 'CAMPEONATO ESPANHOL' in dict1[list1[num]]:
                ws.cell(row, num + 2).fill = PatternFill('solid', fgColor='EEEE00')
            if '- AO VIVO' in dict1[list1[num]] and 'PREMIER LEAGUE' in dict1[list1[num]]:
                ws.cell(row, num + 2).fill = PatternFill('solid', fgColor='9F79EE')
            if '- AO VIVO' in dict1[list1[num]] and 'CAMPEONATO ESPANHOL' in dict1[list1[num]]:
                ws.cell(row, num + 2).fill = PatternFill('solid', fgColor='9F79EE')


if __name__ == '__main__':
    # Driver_Path = r'E:\untitled3\chromedriver\chromedriver.exe'
    # url = 'http://www.espn.com.br/programacao?date=2019-08-17'
    # req = requests.get(url)
    # print(req.text)
    # Selector = etree.HTML(req.text)
    # result = Selector.xpath('//*[@data-channel="espn-brasil"]/div/article[0]')
    # print(result)
    init_web_driver()
    today = datetime.datetime.today()
    # today_str = today.strftime("%Y-%m-%d")
    # print(today_str)
    wb = openpyxl.Workbook()

    for day in range(3):
        date_str = (today + datetime.timedelta(days=day)).strftime("%Y-%m-%d")
        # print(date_str)
        get_data(date_str, wb)
    wb.remove(wb['Sheet'])
    print(wb.sheetnames)
    path = os.getcwd() + r'\espn_epg.xlsx'
    wb.save(path)
    close_web_driver()